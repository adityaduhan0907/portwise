#!/usr/bin/env python3
"""
populate_pb.py  —  OFFLINE batch job (NOT part of the user-facing pipeline).

Fills column G ("P/B") on the 'India Companies' sheet of company_list.xlsx with
each Indian ticker's Price-to-Book, fetched from yfinance.

  * P/B source: ticker.info["priceToBook"]; if missing, fall back to
    currentPrice / bookValue (the same logic module1 uses); else leave BLANK.
  * Throttle-aware: spaces calls out (DELAY) and uses fetch_util.fetch_info
    (retry + exponential backoff) for the rate-limit-prone .info endpoint.
  * Resumable: a row whose G is already filled is SKIPPED, so re-running resumes.
  * Crash-safe: saves periodically via a temp file + atomic replace.
  * Only column G on this ONE sheet is written; C/F and all other sheets/columns
    are left exactly as-is (workbook loaded WITHOUT data_only so nothing else,
    including any formulas, is altered).
"""
import json
import os
import time

import openpyxl

import fetch_util
from fetch_util import TransientFetchError

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
PATH        = os.path.join(SCRIPT_DIR, "company_list.xlsx")
TMP_PATH    = PATH + ".tmp"
SUMMARY     = os.path.join(SCRIPT_DIR, "populate_pb_summary.json")
SHEET_NAME  = "India Companies"
TICKER_COL  = "A"
PB_COL      = "G"
SECTOR_COL  = "C"
MCAP_COL    = "F"
DELAY       = 1.2          # seconds between tickers (throttle-friendly)
SAVE_EVERY  = 25          # rows between periodic atomic saves


def _to_float(x):
    try:
        if x is None:
            return None
        f = float(x)
        return None if f != f else f          # drop NaN
    except (TypeError, ValueError):
        return None


def compute_pb(info):
    """Return (pb_or_None, source) where source in {'priceToBook','computed',''}."""
    pb = _to_float(info.get("priceToBook"))
    if pb is not None:
        return pb, "priceToBook"
    cp = _to_float(info.get("currentPrice"))
    bv = _to_float(info.get("bookValue"))
    if cp is not None and bv not in (None, 0):
        return cp / bv, "computed"
    return None, ""


def atomic_save(wb):
    wb.save(TMP_PATH)
    os.replace(TMP_PATH, PATH)


def main():
    wb = openpyxl.load_workbook(PATH)        # default: preserves formulas/values
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"sheet {SHEET_NAME!r} not found in {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    rows = [r for r in range(2, ws.max_row + 1)
            if ws[f"{TICKER_COL}{r}"].value not in (None, "")]
    total = len(rows)

    n_priceToBook = n_computed = n_blank = n_skipped = 0
    blanks = []

    print(f"[populate_pb] sheet={SHEET_NAME!r} tickers={total} delay={DELAY}s", flush=True)
    processed = 0
    for i, r in enumerate(rows, start=1):
        ticker = str(ws[f"{TICKER_COL}{r}"].value).strip()

        # Resume: skip rows already filled by a prior (partial) run.
        if ws[f"{PB_COL}{r}"].value not in (None, ""):
            n_skipped += 1
            continue

        try:
            info = fetch_util.fetch_info(ticker, use_cache=False)
        except TransientFetchError as exc:
            # Throttle survived all retries -> leave blank (a re-run will retry it).
            print(f"  {i}/{total}  {ticker:<16} THROTTLED (left blank): {exc.detail}", flush=True)
            n_blank += 1
            blanks.append(ticker)
            time.sleep(DELAY * 3)            # back off the whole loop a bit
            continue

        pb, source = compute_pb(info or {})
        if pb is not None:
            ws[f"{PB_COL}{r}"] = round(pb, 2)
            if source == "priceToBook":
                n_priceToBook += 1
            else:
                n_computed += 1
            tag = source
            shown = f"{round(pb,2)}"
        else:
            # leave blank (do NOT write 0) so missing values stay visible
            n_blank += 1
            blanks.append(ticker)
            tag = "blank"
            shown = "-"

        processed += 1
        print(f"  {i}/{total}  {ticker:<16} P/B={shown:<8} [{tag}]", flush=True)

        if processed % SAVE_EVERY == 0:
            atomic_save(wb)
            print(f"  ...saved progress at {i}/{total}", flush=True)

        time.sleep(DELAY)

    atomic_save(wb)

    summary = {
        "sheet": SHEET_NAME, "total_tickers": total,
        "priceToBook": n_priceToBook, "computed": n_computed,
        "blank": n_blank, "skipped_already_filled": n_skipped,
        "blanks": blanks,
    }
    with open(SUMMARY, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)

    print("\n[populate_pb] DONE")
    print(f"  got P/B (priceToBook): {n_priceToBook}")
    print(f"  fell back to computed: {n_computed}")
    print(f"  left blank           : {n_blank}")
    print(f"  skipped (pre-filled) : {n_skipped}")
    print(f"  summary -> {SUMMARY}")


if __name__ == "__main__":
    main()
