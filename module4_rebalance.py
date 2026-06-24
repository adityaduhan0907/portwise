#!/usr/bin/env python3
"""
module4_rebalance.py

Portfolio rebalancing assistant.

Reads:
  - optimised_portfolios.xlsx  (target weights produced by module2/module3)
  - prices.xlsx                (historical prices produced by module1)

Asks the user interactively:
  - Which optimal portfolio to rebalance toward
  - Preferred display currency (USD or INR)
  - Current holdings: ticker symbol + number of shares held

Computes each holding's current market value using the latest price from
prices.xlsx. US stock prices are in USD; Indian stock prices (.NS / .BO)
are in INR and are converted to USD via a live exchange rate.

Generates plain-English BUY / SELL instructions, prints a clean summary,
and exports everything to rebalancing_plan_YYYYMMDD.xlsx.

Run order: module0 -> module1 -> module2 -> module3 -> module4
"""

import os
import sys
import time
import warnings
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import fetch_util

warnings.filterwarnings("ignore")

# ── Configuration ──────────────────────────────────────────────────────────────
SCRIPT_DIR         = os.path.dirname(os.path.abspath(__file__))
THRESHOLD_PCT      = 1.0        # skip trades where |weight diff| <= this percent
PRICES_MAX_AGE_HRS = 24         # warn if prices.xlsx is older than this
W                  = 68         # console output column width

# Tickers with these suffixes have prices quoted in INR on Yahoo Finance
INDIA_SUFFIXES = (".NS", ".BO")

# Summary labels in optimised_portfolios.xlsx that are not ticker rows
SUMMARY_LABELS = {
    "Portfolio Return (%)",
    "Portfolio Volatility (%)",
    "Portfolio Sharpe Ratio",
    "",
    "nan",
}


def _is_summary_label(name):
    """
    True for any non-ticker spacer/summary row: blank/nan, a known summary metric,
    a 'Portfolio ...' label, or anything carrying a '(%)' metric tag (e.g.
    'Portfolio CVaR 95% (%)'). A bare exact-match list misses metrics added later
    (CVaR, VaR, ...); this matches on shape so new footers are rejected too.

    Mirrors module5_report._is_summary_label and the app's parse_goal_sheet filter
    so Module 4, the report and the app all agree on what counts as a ticker.
    """
    s = str(name).strip()
    return (not s) or s.lower() in {x.lower() for x in SUMMARY_LABELS} \
        or s.lower().startswith("portfolio ") or "(%)" in s


# ── Formatting helpers ─────────────────────────────────────────────────────────

def fmt_usd(amount):
    """Return a right-aligned USD string with thousand separators."""
    return f"${amount:>14,.2f}"

def fmt_inr(amount):
    """Return a right-aligned INR string with thousand separators."""
    return f"INR {amount:>14,.2f}"

def fmt_pct(value, signed=False):
    """Return a percentage string with sign if requested."""
    sign = f"{value:+.2f}" if signed else f"{value:.2f}"
    return f"{sign}%"


# ── Freshness check ────────────────────────────────────────────────────────────

def check_prices_freshness(prices_path):
    """
    Compare prices.xlsx modification time to now.
    If older than PRICES_MAX_AGE_HRS, print a prominent warning and return False.
    """
    mtime    = os.path.getmtime(prices_path)
    modified = datetime.fromtimestamp(mtime)
    age_hrs  = (datetime.now() - modified).total_seconds() / 3600

    if age_hrs > PRICES_MAX_AGE_HRS:
        age_str = (f"{age_hrs / 24:.1f} day(s)"
                   if age_hrs >= 48 else f"{age_hrs:.0f} hour(s)")
        bar = "*" * W
        print(f"\n  {bar}")
        print(f"  WARNING: prices.xlsx was last updated {age_str} ago.")
        print(f"           (Last modified: {modified.strftime('%Y-%m-%d %H:%M:%S')})")
        print("  Please rerun module1_data.py before trusting these rebalancing")
        print("  instructions -- prices may not reflect current market values.")
        print(f"  {bar}\n")
        return False

    print(f"  prices.xlsx is current (last modified: {modified.strftime('%Y-%m-%d %H:%M:%S')})")
    return True


# ── Data loaders ───────────────────────────────────────────────────────────────

def load_portfolio_options(xlsx_path):
    """
    Parse optimised_portfolios.xlsx and return all portfolio options.

    Returns
    -------
    dict  {sheet_name: {ticker: weight_pct}}
    Only includes tickers with weight > 0; strips summary rows.
    """
    options = {}
    xf = pd.ExcelFile(xlsx_path)

    for sheet in xf.sheet_names:
        df = pd.read_excel(xlsx_path, sheet_name=sheet)
        weights = {}

        for _, row in df.iterrows():
            stock = str(row.get("Stock", "")).strip()
            wt    = row.get("Weight (%)", None)

            # Skip blank rows and any summary/footer row (Return, Volatility,
            # Sharpe, CVaR, ... -- anything that is not a real ticker).
            if _is_summary_label(stock):
                continue

            try:
                w = float(wt)
                if w > 0:                  # only include non-zero allocations
                    weights[stock] = w
            except (ValueError, TypeError):
                continue

        if weights:
            options[sheet] = weights

    return options


def load_latest_prices(prices_path):
    """
    Read prices.xlsx and return the most recent non-NaN price per ticker.

    Returns
    -------
    dict  {ticker: float}
    """
    df = pd.read_excel(prices_path, index_col=0, parse_dates=True)
    # Forward-fill then take the last row to handle missing data on holidays
    latest = df.ffill().iloc[-1].dropna()
    return latest.to_dict()


def lookup_price(ticker, prices):
    """
    Find a price for the given ticker.
    If an exact match fails and the ticker has no suffix, tries .NS then .BO.

    Returns
    -------
    (resolved_ticker, price)  or  (None, None) if not found.
    """
    if ticker in prices:
        return ticker, prices[ticker]

    # Only try suffix appending if the ticker has no dot already
    if "." not in ticker:
        for suffix in INDIA_SUFFIXES:
            candidate = ticker + suffix
            if candidate in prices:
                print(f"    Auto-resolved '{ticker}' -> '{candidate}'")
                return candidate, prices[candidate]

    return None, None


def is_indian(ticker):
    """Return True if the ticker is NSE or BSE listed (price quoted in INR)."""
    return ticker.upper().endswith(INDIA_SUFFIXES)


# ── Exchange rate ──────────────────────────────────────────────────────────────

def fetch_usd_inr_rate():
    """
    Fetch the live USD/INR spot rate via yfinance (USDINR=X), with up to
    3 retries.  Falls back to manual input if all attempts fail.

    Returns
    -------
    float  — number of INR per USD  (e.g. 95.95)
    """
    print("  Fetching live USD/INR rate (USDINR=X) ...", end=" ", flush=True)
    try:
        hist = fetch_util.fetch_history("USDINR=X", period="3d", what="USD/INR spot")
        if hist is not None and not hist.empty:
            rate = float(hist["Close"].iloc[-1])
            print(f"{rate:.4f} INR/USD")
            return rate
    except fetch_util.TransientFetchError as exc:
        # Auxiliary display-FX rate -> fall back to manual entry rather than abort.
        print(f"failed ({exc.detail})")

    # Manual fallback if yfinance cannot retrieve the rate
    print("  Could not fetch the exchange rate automatically.")
    while True:
        try:
            entry = input("  Enter USD/INR rate manually (e.g. 83.50): ").strip()
            rate  = float(entry)
            if rate > 0:
                return rate
            print("  Rate must be a positive number.")
        except ValueError:
            print("  Please enter a valid number (e.g. 83.50).")


# ── Interactive user input ─────────────────────────────────────────────────────

def get_portfolio_choice(options):
    """
    Display available portfolios and ask the user to choose one.

    Returns
    -------
    (name, weights_dict)
    """
    names = list(options.keys())
    print("\n  Available portfolios from optimised_portfolios.xlsx:")
    print()
    for i, name in enumerate(names, 1):
        tickers = [t for t, w in options[name].items() if w > 0]
        print(f"    {i}.  {name}")
        print(f"         Stocks: {', '.join(tickers)}")
    print()

    while True:
        try:
            choice = int(input("  Enter portfolio number: ").strip())
            if 1 <= choice <= len(names):
                selected = names[choice - 1]
                print(f"  Selected: {selected}")
                return selected, options[selected]
            print(f"  Please enter a number between 1 and {len(names)}.")
        except ValueError:
            print("  Please enter a valid number.")


def get_display_currency():
    """
    Ask whether to show amounts in USD or INR.

    Returns
    -------
    'USD' or 'INR'
    """
    while True:
        resp = input("\n  Display currency — USD or INR? [USD/INR]: ").strip().upper()
        if resp in ("USD", "INR"):
            return resp
        print("  Please type USD or INR.")


def get_holdings(prices):
    """
    Interactively collect the user's current holdings.
    Asks for ticker symbol and number of shares held for each position.
    Warns and skips tickers that cannot be found in prices.xlsx.

    Returns
    -------
    list of (resolved_ticker, n_shares, price_native, is_indian_flag)
    """
    print()
    print("  Enter your current holdings one by one.")
    print("  For each position: type the ticker symbol, press Enter,")
    print("  then enter how many shares you hold.")
    print("  Type 'done' when you have finished entering all positions.")
    print()

    holdings = []

    while True:
        raw = input("  Ticker (or 'done'): ").strip().upper()

        # User signals they are finished
        if raw.lower() == "done":
            if not holdings:
                print("  Please enter at least one holding before typing 'done'.")
                continue
            break

        if not raw:
            continue

        # Find the ticker in prices.xlsx (with .NS / .BO fallback)
        resolved, price = lookup_price(raw, prices)
        if resolved is None:
            print(f"  WARNING: '{raw}' was not found in prices.xlsx -- skipping.")
            print("  Check the ticker symbol and make sure module1 has been run.")
            continue

        # Collect number of shares
        while True:
            try:
                shares_str = (
                    input(f"  Shares of {resolved}: ").strip().replace(",", "")
                )
                n_shares = float(shares_str)
                if n_shares < 0:
                    print("  Number of shares cannot be negative.")
                    continue
                break
            except ValueError:
                print("  Please enter a valid number (e.g. 10 or 10.5).")

        holdings.append((resolved, n_shares, price, is_indian(resolved)))
        print(f"  Added: {resolved}  x  {n_shares:,.4f} shares  @ "
              f"{'INR' if is_indian(resolved) else 'USD'} {price:,.4f}")

    return holdings


# ── Portfolio calculations ─────────────────────────────────────────────────────

def build_current_portfolio(holdings, usd_inr_rate):
    """
    Compute each position's market value in USD and current portfolio weight.

    Indian stocks (.NS / .BO) have prices quoted in INR; divide by usd_inr_rate
    to get the USD equivalent.

    Parameters
    ----------
    holdings     : list of (ticker, n_shares, price_native, is_indian)
    usd_inr_rate : float  — INR per USD (e.g. 95.95)

    Returns
    -------
    positions : dict  {ticker: {shares, price_native, is_indian,
                                value_inr, value_usd, weight_pct}}
    total_usd : float
    """
    positions = {}

    for ticker, n_shares, price_native, indian in holdings:
        if indian:
            # Price is in INR; convert holding value to USD
            value_inr = n_shares * price_native
            value_usd = value_inr / usd_inr_rate
        else:
            # Price is already in USD
            value_inr = None
            value_usd = n_shares * price_native

        positions[ticker] = {
            "shares":       n_shares,
            "price_native": price_native,
            "is_indian":    indian,
            "value_inr":    value_inr,
            "value_usd":    value_usd,
        }

    total_usd = sum(p["value_usd"] for p in positions.values())

    # Calculate each position's share of the total portfolio (in %)
    for data in positions.values():
        data["weight_pct"] = (
            data["value_usd"] / total_usd * 100 if total_usd > 0 else 0.0
        )

    return positions, total_usd


def generate_instructions(positions, target_weights, total_usd, usd_inr_rate):
    """
    Compare current positions to target weights and produce rebalancing trades.

    Rules:
      - Ticker in target only  → BUY (new position, no threshold)
      - Ticker in current only → SELL (exit, no threshold)
      - Both present, |diff| <= THRESHOLD_PCT → skip
      - Both present, |diff|  > THRESHOLD_PCT → BUY or SELL the difference

    Parameters
    ----------
    positions      : dict from build_current_portfolio()
    target_weights : dict {ticker: weight_pct}
    total_usd      : float
    usd_inr_rate   : float

    Returns
    -------
    list of instruction dicts, sorted by |weight_diff| descending
    """
    all_tickers = sorted(set(positions.keys()) | set(target_weights.keys()))
    instructions = []

    for ticker in all_tickers:
        in_current = ticker in positions
        in_target  = ticker in target_weights

        current_w = positions[ticker]["weight_pct"] if in_current else 0.0
        current_v = positions[ticker]["value_usd"]  if in_current else 0.0
        target_w  = target_weights.get(ticker, 0.0)
        target_v  = total_usd * target_w / 100.0

        diff_w   = target_w - current_w        # positive = need to buy
        diff_usd = target_v - current_v        # positive = buy amount, negative = sell

        # Determine the action and whether it is above the threshold
        if not in_target:
            # Position exists but has no place in the optimal portfolio
            action = "SELL"
            amount = current_v                 # sell the entire position
            status = "Recommended"
            note   = "Not in optimal portfolio -- exit full position"

        elif not in_current:
            # Target portfolio requires this stock but the user holds none
            action = "BUY"
            amount = target_v                  # buy the full target allocation
            status = "Recommended"
            note   = "New position"

        elif abs(diff_w) <= THRESHOLD_PCT:
            # Within the 1% rounding tolerance -- not worth trading
            action = "BUY" if diff_w >= 0 else "SELL"
            amount = abs(diff_usd)
            status = (f"Skipped -- diff {diff_w:+.2f}% is within "
                      f"{THRESHOLD_PCT:.0f}% threshold")
            note   = "No trade needed"

        else:
            # Meaningful weight difference -- reweight
            action = "BUY" if diff_w > 0 else "SELL"
            amount = abs(diff_usd)
            status = "Recommended"
            note   = f"Reweight {current_w:.2f}% -> {target_w:.2f}%"

        instructions.append({
            "ticker":     ticker,
            "action":     action,
            "current_w":  current_w,
            "target_w":   target_w,
            "diff_w":     diff_w,
            "diff_usd":   diff_usd,
            "amount_usd": amount,
            "amount_inr": amount * usd_inr_rate,
            "status":     status,
            "note":       note,
        })

    # Sort: recommended trades first (largest diff first), then skipped
    instructions.sort(
        key=lambda x: (0 if "Recommended" in x["status"] else 1, -abs(x["diff_w"]))
    )
    return instructions


# ── Terminal output ────────────────────────────────────────────────────────────

def _sep(char="="):
    return char * W

def print_summary(positions, total_usd, target_name, target_weights,
                  instructions, display_currency, usd_inr_rate, today_str):
    """Print the complete rebalancing plan to the terminal."""

    show_inr  = display_currency == "INR"
    total_inr = total_usd * usd_inr_rate

    # ── Header ─────────────────────────────────────────────────────────────────
    print(f"\n{_sep()}")
    print("  PORTFOLIO REBALANCING PLAN")
    print(f"  Date      : {today_str}")
    print(f"  Target    : {target_name}")
    if show_inr:
        print(f"  USD/INR   : {usd_inr_rate:.4f}")
    print(_sep())

    # ── Total value ────────────────────────────────────────────────────────────
    print(f"\n  Total portfolio value : {fmt_usd(total_usd)}", end="")
    if show_inr:
        print(f"   ({fmt_inr(total_inr)})", end="")
    print()

    # ── Current holdings table ─────────────────────────────────────────────────
    print(f"\n{_sep('-')}")
    print("  CURRENT HOLDINGS")
    print(_sep("-"))

    hdr = (f"  {'Ticker':<20}  {'Shares':>10}  "
           f"{'Price (native)':>16}  {'Native CCY':>10}  "
           f"{'Value USD':>12}  {'Weight':>8}")
    if show_inr:
        hdr += f"  {'Value INR':>14}"
    print(hdr)
    print(f"  {'-'*20}  {'-'*10}  {'-'*16}  {'-'*10}  {'-'*12}  {'-'*8}", end="")
    if show_inr:
        print(f"  {'-'*14}", end="")
    print()

    for ticker, pos in sorted(positions.items()):
        ccy     = "INR" if pos["is_indian"] else "USD"
        inr_val = (pos["value_inr"] if pos["value_inr"] is not None
                   else pos["value_usd"] * usd_inr_rate)
        row = (f"  {ticker:<20}  {pos['shares']:>10,.4f}  "
               f"{pos['price_native']:>16,.4f}  {ccy:>10}  "
               f"${pos['value_usd']:>11,.2f}  {pos['weight_pct']:>7.2f}%")
        if show_inr:
            row += f"  {inr_val:>14,.2f}"
        print(row)

    # ── Target weights table ───────────────────────────────────────────────────
    print(f"\n{_sep('-')}")
    print(f"  TARGET WEIGHTS  ({target_name})")
    print(_sep("-"))

    hdr2 = (f"  {'Ticker':<20}  {'Target %':>9}  "
            f"{'Target Value USD':>18}")
    if show_inr:
        hdr2 += f"  {'Target Value INR':>18}"
    print(hdr2)
    print(f"  {'-'*20}  {'-'*9}  {'-'*18}", end="")
    if show_inr:
        print(f"  {'-'*18}", end="")
    print()

    for ticker, w in sorted(target_weights.items()):
        tv_usd = total_usd * w / 100.0
        row2   = f"  {ticker:<20}  {w:>8.2f}%  ${tv_usd:>17,.2f}"
        if show_inr:
            row2 += f"  {tv_usd * usd_inr_rate:>18,.2f}"
        print(row2)

    # ── Rebalancing instructions ───────────────────────────────────────────────
    print(f"\n{_sep('-')}")
    print("  REBALANCING INSTRUCTIONS")
    print(f"  (Trades with |weight diff| <= {THRESHOLD_PCT:.0f}% are skipped as immaterial)")
    print(_sep("-"))

    actioned = [i for i in instructions if "Recommended" in i["status"]]
    skipped  = [i for i in instructions if "Skipped"     in i["status"]]

    if actioned:
        hdr3 = (f"  {'Action':<5}  {'Ticker':<20}  "
                f"{'Diff %':>8}  {'Amount USD':>14}")
        if show_inr:
            hdr3 += f"  {'Amount INR':>16}"
        hdr3 += "  Note"
        print(hdr3)

        for inst in actioned:
            row3 = (f"  {inst['action']:<5}  {inst['ticker']:<20}  "
                    f"{inst['diff_w']:>+8.2f}%  ${inst['amount_usd']:>13,.2f}")
            if show_inr:
                row3 += f"  {inst['amount_inr']:>16,.2f}"
            row3 += f"  {inst['note']}"
            print(row3)
    else:
        print("  No trades required -- portfolio is already within threshold on all positions.")

    if skipped:
        print(f"\n  Skipped positions (|diff| <= {THRESHOLD_PCT:.0f}%):")
        for inst in skipped:
            print(f"    {inst['ticker']:<22}  diff {inst['diff_w']:+.2f}%  -- {inst['note']}")

    # ── Post-rebalancing estimate ──────────────────────────────────────────────
    # Rebalancing is cash-neutral (sells fund buys), so total value is unchanged.
    print(f"\n{_sep('-')}")
    print(f"  Estimated post-rebalancing value : {fmt_usd(total_usd)}", end="")
    if show_inr:
        print(f"   ({fmt_inr(total_inr)})", end="")
    print()
    print(f"  (Rebalancing is cash-neutral -- no new money added or withdrawn)")
    print(f"\n{_sep()}\n")


# ── Excel export ───────────────────────────────────────────────────────────────

def export_to_excel(positions, total_usd, target_name, target_weights,
                    instructions, display_currency, usd_inr_rate, today_str, out_path):
    """
    Write the rebalancing plan to a formatted Excel workbook with four sheets:
      Summary | Current Holdings | Target Weights | Rebalancing Instructions
    """
    show_inr  = display_currency == "INR"
    total_inr = total_usd * usd_inr_rate

    actioned_count = sum(1 for i in instructions if "Recommended" in i["status"])
    skipped_count  = sum(1 for i in instructions if "Skipped"     in i["status"])

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:

        # ── Sheet 1: Summary ───────────────────────────────────────────────────
        summary_rows = [
            ("Date",                     today_str),
            ("Target Portfolio",         target_name),
            ("Display Currency",         display_currency),
            ("USD/INR Rate",             f"{usd_inr_rate:.4f}" if show_inr else "N/A"),
            ("Total Portfolio (USD)",    round(total_usd, 2)),
            ("Total Portfolio (INR)",    round(total_inr, 2) if show_inr else "N/A"),
            ("Number of Holdings",       len(positions)),
            ("Trades Required",          actioned_count),
            ("Trades Skipped (< 1%)",    skipped_count),
        ]
        pd.DataFrame(summary_rows, columns=["Field", "Value"]).to_excel(
            writer, sheet_name="Summary", index=False
        )

        # ── Sheet 2: Current Holdings ──────────────────────────────────────────
        ch_rows = []
        for ticker, pos in sorted(positions.items()):
            inr_val = (pos["value_inr"] if pos["value_inr"] is not None
                       else pos["value_usd"] * usd_inr_rate)
            row = {
                "Ticker":            ticker,
                "Shares":            round(pos["shares"], 4),
                "Price (native)":    round(pos["price_native"], 4),
                "Native Currency":   "INR" if pos["is_indian"] else "USD",
                "Value (USD)":       round(pos["value_usd"], 2),
                "Current Weight %":  round(pos["weight_pct"], 2),
            }
            if show_inr:
                row["Value (INR)"] = round(inr_val, 2)
            ch_rows.append(row)
        pd.DataFrame(ch_rows).to_excel(
            writer, sheet_name="Current Holdings", index=False
        )

        # ── Sheet 3: Target Weights ────────────────────────────────────────────
        tw_rows = []
        for ticker, w in sorted(target_weights.items()):
            tv_usd = total_usd * w / 100.0
            row = {
                "Ticker":              ticker,
                "Target Weight %":     round(w, 2),
                "Target Value (USD)":  round(tv_usd, 2),
            }
            if show_inr:
                row["Target Value (INR)"] = round(tv_usd * usd_inr_rate, 2)
            tw_rows.append(row)
        pd.DataFrame(tw_rows).to_excel(
            writer, sheet_name="Target Weights", index=False
        )

        # ── Sheet 4: Rebalancing Instructions ─────────────────────────────────
        ri_rows = []
        for inst in instructions:
            row = {
                "Action":               inst["action"],
                "Ticker":               inst["ticker"],
                "Current Weight %":     round(inst["current_w"], 2),
                "Target Weight %":      round(inst["target_w"],  2),
                "Weight Difference %":  round(inst["diff_w"],    2),
                "Amount (USD)":         round(inst["amount_usd"], 2),
                "Status":               inst["status"],
                "Note":                 inst["note"],
            }
            if show_inr:
                row["Amount (INR)"] = round(inst["amount_inr"], 2)
            ri_rows.append(row)
        pd.DataFrame(ri_rows).to_excel(
            writer, sheet_name="Rebalancing Instructions", index=False
        )

        # ── Apply openpyxl formatting to all sheets ────────────────────────────
        _style_workbook(writer.sheets, show_inr)


def _style_workbook(sheets, show_inr):
    """
    Apply consistent formatting to all sheets in the workbook:
      - Dark blue header row with white bold text
      - BUY rows: light green; SELL rows: light red; Skipped rows: light grey
      - Auto column widths (capped at 40)
      - Number formats for currency and percentage columns
    """
    HDR_FONT  = Font(bold=True, color="FFFFFF")
    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    BUY_FILL  = PatternFill("solid", fgColor="C6EFCE")   # green -- supplementary
    SELL_FILL = PatternFill("solid", fgColor="FFC7CE")   # red   -- supplementary
    SKIP_FILL = PatternFill("solid", fgColor="EEEEEE")   # grey

    # Number formats applied by column header keyword
    USD_FMT  = '#,##0.00'
    INR_FMT  = '#,##0.00'
    PCT_FMT  = '0.00'

    for ws in sheets.values():
        # Style the header row
        for cell in ws[1]:
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Identify column positions by their header label
        col_meta = {}          # col_letter -> header label
        action_col = None
        status_col = None
        for cell in ws[1]:
            label = str(cell.value or "")
            col_meta[cell.column] = label
            if label == "Action":
                action_col = cell.column
            if label == "Status":
                status_col = cell.column

        # Apply number formats and row colour coding for data rows
        for row in ws.iter_rows(min_row=2):
            action_val = (row[action_col - 1].value
                          if action_col and len(row) >= action_col else "")
            status_val = str(row[status_col - 1].value
                             if status_col and len(row) >= status_col else "")

            # Row fill: NOTE colour is supplementary — text (BUY/SELL/Skipped) is primary
            if "Skipped" in status_val:
                row_fill = SKIP_FILL
            elif action_val == "BUY":
                row_fill = BUY_FILL
            elif action_val == "SELL":
                row_fill = SELL_FILL
            else:
                row_fill = None

            for cell in row:
                if row_fill:
                    cell.fill = row_fill

                # Apply number formats by column header
                label = col_meta.get(cell.column, "")
                if "USD" in label or "Value" in label and "INR" not in label:
                    cell.number_format = USD_FMT
                elif "INR" in label:
                    cell.number_format = INR_FMT
                elif "%" in label or "Weight" in label:
                    cell.number_format = PCT_FMT
                elif "Shares" in label or "Price" in label:
                    cell.number_format = '#,##0.0000'

        # Auto-size columns (capped at 40 characters wide)
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=8,
            )
            ws.column_dimensions[
                get_column_letter(col[0].column)
            ].width = min(max_len + 3, 40)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    today_str  = datetime.now().strftime("%Y-%m-%d")
    date_stamp = datetime.now().strftime("%Y%m%d")

    print(f"\n{'='*W}")
    print("  Module 4 -- Portfolio Rebalancing Assistant")
    print(f"  {today_str}")
    print(f"{'='*W}\n")

    # ── Verify required input files exist ──────────────────────────────────────
    portfolios_path = os.path.join(SCRIPT_DIR, "optimised_portfolios.xlsx")
    prices_path     = os.path.join(SCRIPT_DIR, "prices.xlsx")

    for fpath, fname in [(portfolios_path, "optimised_portfolios.xlsx"),
                         (prices_path,     "prices.xlsx")]:
        if not os.path.exists(fpath):
            print(f"  ERROR: '{fname}' not found in {SCRIPT_DIR}.")
            print("  Run the preceding modules in order:")
            print("    module0 -> module1 -> module2 -> module3 -> module4")
            sys.exit(1)

    # ── Check price data freshness ─────────────────────────────────────────────
    check_prices_freshness(prices_path)

    # ── Load the portfolio options ─────────────────────────────────────────────
    try:
        portfolio_options = load_portfolio_options(portfolios_path)
    except Exception as exc:
        print(f"  ERROR reading optimised_portfolios.xlsx: {exc}")
        sys.exit(1)

    if not portfolio_options:
        print("  ERROR: No valid portfolios found in optimised_portfolios.xlsx.")
        print("  Please rerun module2_optimiser.py to regenerate the file.")
        sys.exit(1)

    # ── Load the latest prices ─────────────────────────────────────────────────
    try:
        prices = load_latest_prices(prices_path)
    except Exception as exc:
        print(f"  ERROR reading prices.xlsx: {exc}")
        sys.exit(1)

    print(f"  Loaded {len(prices)} latest prices from prices.xlsx.")

    # ── User choices: portfolio, display currency, exchange rate ───────────────
    target_name, target_weights = get_portfolio_choice(portfolio_options)
    display_currency             = get_display_currency()

    # Always fetch the exchange rate: needed to convert Indian stock prices
    # (quoted in INR in prices.xlsx) to USD for unified calculations.
    print()
    usd_inr_rate = fetch_usd_inr_rate()

    # ── Collect current holdings ───────────────────────────────────────────────
    raw_holdings = get_holdings(prices)

    if not raw_holdings:
        print("  No valid holdings entered. Exiting.")
        sys.exit(0)

    # ── Calculate current portfolio values and weights ─────────────────────────
    positions, total_usd = build_current_portfolio(raw_holdings, usd_inr_rate)

    if total_usd <= 0:
        print("  ERROR: Total portfolio value is zero -- cannot calculate weights.")
        sys.exit(1)

    # ── Generate rebalancing instructions ──────────────────────────────────────
    instructions = generate_instructions(
        positions, target_weights, total_usd, usd_inr_rate
    )

    # ── Print the full plan to the terminal ────────────────────────────────────
    print_summary(
        positions, total_usd, target_name, target_weights,
        instructions, display_currency, usd_inr_rate, today_str,
    )

    # ── Export to Excel ────────────────────────────────────────────────────────
    out_path = os.path.join(SCRIPT_DIR, f"rebalancing_plan_{date_stamp}.xlsx")
    try:
        export_to_excel(
            positions, total_usd, target_name, target_weights,
            instructions, display_currency, usd_inr_rate, today_str, out_path,
        )
    except Exception as exc:
        print(f"  ERROR exporting to Excel: {exc}")
        sys.exit(1)

    print(f"\n{'='*W}\n")


if __name__ == "__main__":
    main()
