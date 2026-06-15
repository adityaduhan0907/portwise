#!/usr/bin/env python3
"""
module0_riskfree.py

Run this FIRST before any other module.

Risk-free rate sources:
  - US  : read from data_points.xlsx cell B6 (percent -> decimal).
          [PAUSED: live 13-week T-bill yield via ^IRX (yfinance)]
  - India: 10-year G-Sec yield via IN10Y-IN.BO  (yfinance)
            Falls back to 6.5 % if the ticker is unavailable.

Calculates a blended rate weighted by the proportion of US vs Indian
stocks in the portfolio, then writes all three rates to risk_free_rates.json
so module1, module2, and module3 can read from one shared source of truth.
"""

import json
import os
import sys
import time
import warnings
from datetime import datetime

import numpy as np
import pandas as pd
import yfinance as yf

warnings.filterwarnings("ignore")

# ── Configuration ──────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_PATH  = os.path.join(SCRIPT_DIR, "risk_free_rates.json")
DATA_POINTS_PATH = os.path.join(SCRIPT_DIR, "data_points.xlsx")  # US RF source (cell B6)

# ── Factor-history ingestion (Layer 1 → module 1 handoff) ───────────────────────
# Fama-French monthly factor history. Sheets are labelled by their literal name
# (the file stores them out of order), and columns are selected BY NAME because
# the two markets list factors in a different order.
FACTOR_PATH      = os.path.join(SCRIPT_DIR, "Fama_French_Factors.xlsx")
FACTOR_JSON_PATH = os.path.join(SCRIPT_DIR, "factor_history.json")  # module1 handoff

# market -> sheet name + factor columns to keep (BY NAME) + earliest year to keep.
# RF is always kept in addition to these factors.
FACTOR_SHEETS = {
    "India": {"sheet": "Sheet1", "factors": ["MF", "SMB", "HML", "WML"], "min_year": 1993},
    "USA":   {"sheet": "Sheet2", "factors": ["Mkt-RF", "SMB", "HML"],     "min_year": 1926},
}
KF_SENTINELS       = [-99.99, -999.0, -999.99]  # Ken-French missing-data markers
MIN_HISTORY_MONTHS = 36   # assets with fewer monthly obs trigger factor covariance

# Default fallback rates (decimal form)
US_FALLBACK_RATE    = 0.043   # 4.3 %  — approximate 2024/25 T-bill rate
INDIA_FALLBACK_RATE = 0.065   # 6.5 %  — approximate 10-year G-Sec rate

# yfinance ticker symbols
US_TICKER     = "^IRX"          # 13-week Treasury bill (annualised %)
INDIA_TICKER  = "IN10Y-IN.BO"   # India 10-year G-Sec proxy

def _load_portfolio_tickers():
    """
    Return the ticker list for US/India classification.
    When called from run_all.py, reads the user's tickers from run_config.json.
    Falls back to the hardcoded default when run standalone.
    """
    try:
        cfg_path = os.path.join(SCRIPT_DIR, "run_config.json")
        with open(cfg_path, encoding="utf-8") as f:
            cfg = json.load(f)
        tickers = cfg.get("tickers")
        if tickers:
            return tickers
    except FileNotFoundError:
        pass
    except Exception as exc:
        print(f"  WARNING: Could not read run_config.json for tickers ({exc}).")
    # Default standalone ticker list (edit here for standalone use)
    return [
        "AAPL", "MSFT", "GOOGL",
        "RELIANCE.NS", "TCS.NS", "INFY.NS", "HDFCBANK.NS", "WIPRO.BO",
    ]

PORTFOLIO_TICKERS = _load_portfolio_tickers()

# ── Fetch helpers ──────────────────────────────────────────────────────────────

def fetch_latest_close(symbol, lookback="10d", max_retries=3, delay=2):
    """
    Return the most recent closing value for a yfinance ticker, or None on
    any failure (no data, HTTP error, empty frame, etc.).  Retries up to
    max_retries times before giving up.
    """
    for attempt in range(max_retries):
        try:
            hist = yf.Ticker(symbol).history(period=lookback, auto_adjust=True)
            if isinstance(hist.columns, object) and "Close" in hist.columns:
                close = hist["Close"].dropna()
                if not close.empty:
                    return float(close.iloc[-1])
            return None
        except Exception:
            if attempt < max_retries - 1:
                time.sleep(delay)
    return None


# === PAUSED (re-enable later): live US risk-free rate via ^TNX ===
# NOTE: the live US fetch actually uses ^IRX (13-week T-bill), not ^TNX.
# To restore: uncomment this function and switch the call site in main()
# back from read_us_rate_from_file() to fetch_us_rate().
# def fetch_us_rate():
#     """
#     Fetch the annualised 13-week T-bill yield via ^IRX.
#     ^IRX is quoted as a percentage (e.g. 3.59 means 3.59 %), so divide by 100.
#     Returns (rate_decimal, source_description).
#     """
#     raw = fetch_latest_close(US_TICKER)
#     if raw is not None and raw > 0:
#         return raw / 100.0, f"{US_TICKER} (13-week T-bill, last close {raw:.3f}%)"
#
#     print(f"  WARNING: Could not fetch {US_TICKER}. Using fallback {US_FALLBACK_RATE*100:.2f}%.")
#     return US_FALLBACK_RATE, f"fallback ({US_TICKER} unavailable)"
# === END PAUSED ===


def read_rates_from_file():
    """
    Read both risk-free rates from data_points.xlsx in a single open:
      - US    : cell B6 (percent -> decimal), e.g. 3.241 -> 0.03241
      - India : cell B7 (percent -> decimal), e.g. 6.71  -> 0.0671
    Returns (us_rate, us_source, india_rate, india_source).  Each side falls
    back to its *_FALLBACK_RATE independently if that cell is missing/invalid.
    """
    us_rate,    us_source    = US_FALLBACK_RATE,    "fallback (data_points.xlsx!B6 unavailable)"
    india_rate, india_source = INDIA_FALLBACK_RATE, "fallback (data_points.xlsx!B7 unavailable)"

    try:
        import openpyxl
        wb = openpyxl.load_workbook(DATA_POINTS_PATH, data_only=True)
        ws = wb.active
        raw_us    = ws["B6"].value
        raw_india = ws["B7"].value
        wb.close()

        if raw_us is not None and float(raw_us) > 0:
            us_rate, us_source = float(raw_us) / 100.0, f"data_points.xlsx!B6 ({float(raw_us):.3f}%)"
        else:
            print("  WARNING: data_points.xlsx!B6 is empty or non-positive.")

        if raw_india is not None and float(raw_india) > 0:
            india_rate, india_source = float(raw_india) / 100.0, f"data_points.xlsx!B7 ({float(raw_india):.3f}%)"
        else:
            print("  WARNING: data_points.xlsx!B7 is empty or non-positive.")
    except Exception as exc:
        print(f"  WARNING: Could not read rates from data_points.xlsx ({exc}).")

    if us_source.startswith("fallback"):
        print(f"  WARNING: Using fallback US risk-free rate {US_FALLBACK_RATE*100:.2f}%.")
    if india_source.startswith("fallback"):
        print(f"  WARNING: Using fallback Indian risk-free rate {INDIA_FALLBACK_RATE*100:.2f}%.")

    return us_rate, us_source, india_rate, india_source


def fetch_india_rate():
    """
    Fetch the India 10-year G-Sec yield via IN10Y-IN.BO.
    Quoted as a percentage on Yahoo Finance; divide by 100.
    Falls back to INDIA_FALLBACK_RATE if the ticker is unavailable.
    Returns (rate_decimal, source_description).
    """
    raw = fetch_latest_close(INDIA_TICKER)
    if raw is not None and raw > 0:
        return raw / 100.0, f"{INDIA_TICKER} (India 10-year G-Sec, last close {raw:.3f}%)"

    print(f"  WARNING: Could not fetch {INDIA_TICKER}.")
    print(f"  WARNING: Using fallback Indian risk-free rate of {INDIA_FALLBACK_RATE*100:.2f}%.")
    return INDIA_FALLBACK_RATE, f"fallback ({INDIA_TICKER} unavailable, default {INDIA_FALLBACK_RATE*100:.1f}%)"


# ── Factor-history ingestion + cleaning ─────────────────────────────────────────

def _parse_factor_date(val):
    """
    Parse a Ken-French style date into a Timestamp (month resolution).
    Handles India 'YYYY-MM' strings and US integer 'YYYYMM' (and YYYYMMDD).
    Returns a pd.Timestamp or pd.NaT on failure.
    """
    if val is None:
        return pd.NaT
    s = str(val).strip()
    if s == "" or s.lower() in ("na", "nan", "none", "#value!"):
        return pd.NaT
    try:
        if "-" in s or "/" in s:
            return pd.to_datetime(s, errors="coerce")
        digits = s.split(".")[0]            # tolerate '192607.0'
        if len(digits) == 6:
            return pd.to_datetime(digits, format="%Y%m", errors="coerce")
        if len(digits) == 8:
            return pd.to_datetime(digits, format="%Y%m%d", errors="coerce")
        return pd.to_datetime(s, errors="coerce")
    except Exception:
        return pd.NaT


def _clean_market_factors(market, spec, path=FACTOR_PATH):
    """
    Read and clean one market's monthly factor history.

    Steps: select Date + named factor columns + RF -> coerce numeric (bad strings
    like 'NA'/'#VALUE!' become NaN) -> convert Ken-French sentinels to NaN ->
    parse Date -> drop any row that fails to parse or has a missing value ->
    keep only rows on/after the market's earliest year -> convert percent to
    decimal monthly returns.

    Returns a cleaned DataFrame indexed by month-end Timestamp with columns
    [<factors...>, 'RF'] (decimals), or an empty DataFrame on failure.
    """
    raw = pd.read_excel(path, sheet_name=spec["sheet"])
    date_col = raw.columns[0]                       # India 'Date' / US 'time'

    keep_factors = [c for c in spec["factors"] if c in raw.columns]
    val_cols     = keep_factors + (["RF"] if "RF" in raw.columns else [])

    df = raw[[date_col] + val_cols].copy()
    df["Date"] = df[date_col].apply(_parse_factor_date)

    for c in val_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce")        # 'NA'/'#VALUE!' -> NaN
        df[c] = df[c].replace(KF_SENTINELS, np.nan)          # -99.99 / -999 -> NaN

    df = df.dropna(subset=["Date"] + val_cols)               # drop unparsable / missing
    df = df[df["Date"].dt.year >= spec["min_year"]]
    df = df.sort_values("Date").set_index("Date")

    df = df[val_cols] / 100.0                                 # percent -> decimal
    return df


def ingest_factor_history(path=FACTOR_PATH):
    """
    Ingest and clean factor history for every configured market.
    Returns {market: cleaned DataFrame}. Markets that fail to load are skipped
    with a warning so the rest of module 0 still runs.
    """
    markets = {}
    if not os.path.exists(path):
        print(f"  WARNING: factor history file not found ({path}). "
              "Skipping factor ingestion.")
        return markets

    for market, spec in FACTOR_SHEETS.items():
        try:
            df = _clean_market_factors(market, spec, path)
            if df.empty:
                print(f"  WARNING: {market} factor sheet '{spec['sheet']}' "
                      "cleaned to zero rows.")
            markets[market] = df
        except Exception as exc:
            print(f"  WARNING: Could not ingest {market} factors "
                  f"(sheet '{spec['sheet']}'): {exc}")
    return markets


# ── Per-asset history (36-month covariance flag) ─────────────────────────────────

def _monthly_history_count(raw_ticker, start="1990-01-01"):
    """
    Return (resolved_symbol, n_months) for one portfolio asset by downloading
    monthly closes. Mirrors module1's resolution (try as-is, then .NS, .BO).
    On total failure returns (None, 0).
    """
    candidates = (
        [raw_ticker] if "." in raw_ticker
        else [raw_ticker, f"{raw_ticker}.NS", f"{raw_ticker}.BO"]
    )
    for symbol in candidates:
        try:
            hist = yf.download(
                symbol, start=start, interval="1mo",
                auto_adjust=True, progress=False, threads=False,
            )
            if hist is None or hist.empty:
                continue
            close = hist["Close"]
            if isinstance(close, pd.DataFrame):
                close = close.iloc[:, 0]
            n = int(close.dropna().shape[0])
            if n > 0:
                return symbol, n
        except Exception:
            continue
    return None, 0


def assess_asset_history(tickers):
    """
    Count months of monthly history for each portfolio asset and decide whether
    module 1 should fall back to a factor-based covariance.

    If ANY asset has < MIN_HISTORY_MONTHS observations (including assets whose
    history could not be fetched), use_factor_covariance is True.

    Returns (history_dict, use_factor_covariance) where history_dict maps the
    original ticker -> {"resolved", "months", "ok"}.
    """
    history = {}
    use_factor_cov = False
    for raw in tickers:
        symbol, n = _monthly_history_count(raw)
        short = (n < MIN_HISTORY_MONTHS)          # 0 (fetch failed) counts as short
        if short:
            use_factor_cov = True
        history[raw] = {
            "resolved": symbol,
            "months":   n,
            "ok":       symbol is not None,
        }
    return history, use_factor_cov


def write_factor_history(markets, asset_history, use_factor_cov, generated_at):
    """
    Write cleaned factor returns, monthly RF, per-asset history and the
    covariance flag to factor_history.json — the same write-a-JSON handoff
    pattern module 0 already uses for risk_free_rates.json. Module 1 reads this
    file; risk_free_rates.json is left untouched.
    """
    market_payload = {}
    for market, df in markets.items():
        if df is None or df.empty:
            continue
        factor_cols = [c for c in df.columns if c != "RF"]
        dates = [d.strftime("%Y-%m") for d in df.index]
        market_payload[market] = {
            "date_start":     dates[0] if dates else None,
            "date_end":       dates[-1] if dates else None,
            "n_months":       len(df),
            "columns":        list(df.columns),     # factors + RF (by name)
            "units":          "decimal monthly returns",
            "dates":          dates,
            "factor_returns": {c: [round(float(x), 8) for x in df[c]] for c in factor_cols},
            "monthly_rf":     [round(float(x), 8) for x in df["RF"]] if "RF" in df.columns else [],
        }

    payload = {
        "generated_at":              generated_at,
        "use_factor_covariance":     bool(use_factor_cov),
        "min_history_months":        MIN_HISTORY_MONTHS,
        "asset_history": {
            t: {"resolved": v["resolved"], "months": v["months"]}
            for t, v in asset_history.items()
        },
        "markets": market_payload,
    }

    with open(FACTOR_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    return FACTOR_JSON_PATH


# ── Portfolio classification ───────────────────────────────────────────────────

def classify_tickers(tickers):
    """
    Split tickers into Indian (.NS or .BO suffix) and US (everything else).
    Returns (us_list, india_list).
    """
    india = [t for t in tickers if t.upper().endswith(".NS") or t.upper().endswith(".BO")]
    us    = [t for t in tickers if t not in india]
    return us, india


def blended_rate(us_rate, india_rate, us_tickers, india_tickers):
    """
    Weighted average of US and Indian rates by stock count.
    Returns (blended_decimal, us_weight, india_weight).
    """
    total = len(us_tickers) + len(india_tickers)
    if total == 0:
        return us_rate, 1.0, 0.0

    w_us    = len(us_tickers)    / total
    w_india = len(india_tickers) / total
    blended = us_rate * w_us + india_rate * w_india
    return blended, w_us, w_india


# ── Display helpers ────────────────────────────────────────────────────────────

def print_section(title, width=62):
    print(f"\n{'='*width}")
    print(f"  {title}")
    print(f"{'='*width}")


def print_rate_line(label, rate_decimal, source):
    print(f"  {label:<30} {rate_decimal*100:>6.4f}%  ({source})")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    fetched_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    print(f"\n{'='*62}")
    print("  Module 0 -- Risk-Free Rate Fetcher")
    print(f"  Rates fetched at : {fetched_at}")
    print(f"{'='*62}")

    # ── 1. Fetch live rates ────────────────────────────────────────────────────
    print_section("FETCHING RATES")

    us_rate, us_source, india_rate, india_source = read_rates_from_file()

    print_rate_line("US risk-free rate    ", us_rate,    us_source)
    print_rate_line("India risk-free rate ", india_rate, india_source)

    # ── 2. Classify portfolio tickers ─────────────────────────────────────────
    print_section("PORTFOLIO CLASSIFICATION")

    us_tickers, india_tickers = classify_tickers(PORTFOLIO_TICKERS)
    total = len(us_tickers) + len(india_tickers)

    print(f"  Total tickers    : {total}")
    print(f"  US stocks  ({len(us_tickers):>2}) : {', '.join(us_tickers) if us_tickers else 'none'}")
    print(f"  Indian stocks ({len(india_tickers):>2}): {', '.join(india_tickers) if india_tickers else 'none'}")

    # ── 3. Blended rate ────────────────────────────────────────────────────────
    print_section("BLENDED RATE CALCULATION")

    blended, w_us, w_india = blended_rate(us_rate, india_rate, us_tickers, india_tickers)

    print(f"  US weight        : {w_us*100:.1f}%  ({len(us_tickers)} of {total} stocks)")
    print(f"  India weight     : {w_india*100:.1f}%  ({len(india_tickers)} of {total} stocks)")
    print()
    print(f"  Blended = ({us_rate*100:.4f}% x {w_us*100:.1f}%) + ({india_rate*100:.4f}% x {w_india*100:.1f}%)")
    print(f"          = {us_rate*100*w_us:.4f}% + {india_rate*100*w_india:.4f}%")
    print(f"          = {blended*100:.4f}%")

    # ── 4. Summary ─────────────────────────────────────────────────────────────
    print_section("RATE SUMMARY")
    width_label = 30
    print(f"  {'US Risk-Free Rate':<{width_label}} {us_rate*100:>8.4f}%")
    print(f"  {'Indian Risk-Free Rate':<{width_label}} {india_rate*100:>8.4f}%")
    print(f"  {'Blended Risk-Free Rate':<{width_label}} {blended*100:>8.4f}%")
    print(f"\n  Use BLENDED rate in module1/2/3 for this mixed portfolio.")

    # ── 4b. Factor history + 36-month covariance flag (module1 handoff) ─────────
    print_section("FACTOR HISTORY & COVARIANCE FLAG")

    markets = ingest_factor_history()
    for market in FACTOR_SHEETS:
        df = markets.get(market)
        if df is None or df.empty:
            print(f"  {market:<6}: no usable factor rows.")
            continue
        cols = ", ".join(df.columns)            # includes RF, labelled by name
        print(f"  {market:<6}: {df.index[0].strftime('%Y-%m')} -> "
              f"{df.index[-1].strftime('%Y-%m')}   "
              f"{len(df):>4} monthly rows   cols: [{cols}]")

    print()
    print(f"  Per-asset months of history (threshold = {MIN_HISTORY_MONTHS}):")
    asset_history, use_factor_cov = assess_asset_history(PORTFOLIO_TICKERS)
    for raw, info in asset_history.items():
        resolved = info["resolved"] or "unresolved"
        flag     = "  << short" if info["months"] < MIN_HISTORY_MONTHS else ""
        label    = raw if resolved in (raw, None) else f"{raw} -> {resolved}"
        print(f"    {label:<26} {info['months']:>4} months{flag}")

    print()
    print(f"  use_factor_covariance = {use_factor_cov}   "
          f"({'a short-history asset was found' if use_factor_cov else 'all assets have >= 36 months'})")

    try:
        path = write_factor_history(markets, asset_history, use_factor_cov, fetched_at)
        print(f"  Saved -> {path}")
    except Exception as exc:
        print(f"  ERROR: Could not write {FACTOR_JSON_PATH}: {exc}")

    # ── 5. Write JSON ──────────────────────────────────────────────────────────
    payload = {
        "fetched_at":        fetched_at,
        "us_rate":           round(us_rate,    8),
        "india_rate":        round(india_rate, 8),
        "blended_rate":      round(blended,    8),
        "us_rate_pct":       round(us_rate    * 100, 4),
        "india_rate_pct":    round(india_rate * 100, 4),
        "blended_rate_pct":  round(blended    * 100, 4),
        "us_rate_source":    us_source,
        "india_rate_source": india_source,
        "portfolio": {
            "total_stocks":      total,
            "us_stocks":         us_tickers,
            "india_stocks":      india_tickers,
            "us_proportion":     round(w_us,    6),
            "india_proportion":  round(w_india, 6),
        },
    }

    try:
        with open(JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2)
        print(f"\n  Saved -> {JSON_PATH}")
    except Exception as exc:
        print(f"\n  ERROR: Could not write {JSON_PATH}: {exc}")
        sys.exit(1)

    print(f"\n{'='*62}\n")


if __name__ == "__main__":
    main()
