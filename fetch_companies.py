

import sys
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "company_list.xlsx"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; portfolio-research-tool/1.0)"
}

INDICES = [
    {
        "name": "SP500",
        "label": "S&P 500",
        "url": "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies",
        "table_index": 0,
        "col_map": {"Symbol": "Ticker", "Security": "Company", "GICS Sector": "Sector"},
    },
    {
        "name": "SP400",
        "label": "S&P 400",
        "url": "https://en.wikipedia.org/wiki/List_of_S%26P_400_companies",
        "table_index": 0,
        "col_map": {"Symbol": "Ticker", "Security": "Company", "GICS Sector": "Sector"},
    },
    {
        "name": "SP600",
        "label": "S&P 600",
        "url": "https://en.wikipedia.org/wiki/List_of_S%26P_600_companies",
        "table_index": 0,
        "col_map": {"Symbol": "Ticker", "Security": "Company", "GICS Sector": "Sector"},
    },
]


def clean_ticker(ticker: str) -> str:
    return str(ticker).strip().upper().replace(".", "-")


def fetch_index(info: dict, session: requests.Session) -> pd.DataFrame | None:
    print(f"Fetching {info['label']} companies from Wikipedia...")
    try:
        tables = pd.read_html(
            info["url"],
            storage_options={"User-Agent": HEADERS["User-Agent"]},
        )
        df = tables[info["table_index"]]

        # Flatten multi-level columns if present
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = [" ".join(str(c) for c in col).strip() for col in df.columns]

        # Attempt to map known column names; fall back to positional
        col_map = info["col_map"]
        matched = {wiki_col: out_col for wiki_col, out_col in col_map.items() if wiki_col in df.columns}

        if len(matched) < 2:
            # Try case-insensitive match
            lower_cols = {c.lower(): c for c in df.columns}
            matched = {}
            for wiki_col, out_col in col_map.items():
                if wiki_col.lower() in lower_cols:
                    matched[lower_cols[wiki_col.lower()]] = out_col

        if not matched:
            print(f"  WARNING: Could not map columns for {info['label']}. Available: {list(df.columns)}")
            return None

        df = df.rename(columns=matched)

        # Keep only the three target columns that were successfully mapped
        keep = [c for c in ["Ticker", "Company", "Sector"] if c in df.columns]
        df = df[keep].copy()

        # Fill missing columns
        for col in ["Ticker", "Company", "Sector"]:
            if col not in df.columns:
                df[col] = ""

        df["Ticker"] = df["Ticker"].apply(clean_ticker)
        df["Company"] = df["Company"].astype(str).str.strip()
        df["Sector"] = df["Sector"].astype(str).str.strip()
        df["Index"] = info["name"]

        # Drop rows where ticker looks like a header or is empty
        df = df[df["Ticker"].str.len() > 0]
        df = df[df["Ticker"] != "TICKER"]

        count = len(df)
        print(f"[OK] {info['label']}: {count:,} companies fetched")
        return df

    except Exception as exc:
        print(f"WARNING: Could not fetch {info['label']}. Continuing with remaining indices.")
        print(f"  Detail: {exc}")
        return None


def format_workbook(path: str, summary: dict) -> None:
    wb = load_workbook(path)

    # ── Sheet 1: US Companies ──────────────────────────────────────────────
    ws = wb["US Companies"]

    navy_fill = PatternFill(fill_type="solid", fgColor="1B3A6B")
    white_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    # ── Sheet 2: Summary ───────────────────────────────────────────────────
    ws2 = wb["Summary"]

    header_fill = PatternFill(fill_type="solid", fgColor="1B3A6B")
    header_font = Font(bold=True, color="FFFFFF")
    label_font = Font(bold=True)

    ws2["A1"] = "Index"
    ws2["B1"] = "Companies"
    for cell in [ws2["A1"], ws2["B1"]]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    rows = [
        ("S&P 500", summary.get("SP500", 0)),
        ("S&P 400", summary.get("SP400", 0)),
        ("S&P 600", summary.get("SP600", 0)),
        ("Total", summary.get("Total", 0)),
    ]
    for r_idx, (label, count) in enumerate(rows, start=2):
        ws2.cell(row=r_idx, column=1, value=label).font = label_font
        ws2.cell(row=r_idx, column=2, value=count).alignment = Alignment(horizontal="center")

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 14

    wb.save(path)


def main() -> None:
    session = requests.Session()
    session.headers.update(HEADERS)

    frames = []
    for info in INDICES:
        df = fetch_index(info, session)
        if df is not None:
            frames.append(df)

    if not frames:
        print("ERROR: Could not fetch any company data. Please check your internet connection.")
        sys.exit(1)

    print("\nCombining and cleaning data...")
    combined = pd.concat(frames, ignore_index=True)
    combined = combined.drop_duplicates(subset="Ticker", keep="first")
    combined = combined.sort_values("Company", ignore_index=True)
    combined = combined[["Ticker", "Company", "Sector", "Index"]]
    print(f"[OK] Total: {len(combined):,} unique companies")

    # Build per-index counts before writing
    summary = combined["Index"].value_counts().to_dict()
    summary["Total"] = len(combined)

    print(f"\nSaving to {OUTPUT_FILE}...")
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        combined.to_excel(writer, sheet_name="US Companies", index=False)

        summary_rows = [
            ["S&P 500", summary.get("SP500", 0)],
            ["S&P 400", summary.get("SP400", 0)],
            ["S&P 600", summary.get("SP600", 0)],
            ["Total", summary["Total"]],
        ]
        pd.DataFrame(summary_rows, columns=["Index", "Companies"]).to_excel(
            writer, sheet_name="Summary", index=False
        )

    format_workbook(OUTPUT_FILE, summary)
    print(f"[OK] Saved to {OUTPUT_FILE}")
    print(f"\nColumn names in 'US Companies' sheet: Ticker | Company | Sector | Index")


if __name__ == "__main__":
    main()
