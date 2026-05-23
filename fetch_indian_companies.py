




























import sys
import os
import re
import random
import pandas as pd

# Force UTF-8 output so Unicode chars (₹, etc.) don't crash on Windows terminals
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import yfinance as yf
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_FILE = r"C:\Users\adity\Documents\FinanceScripts\company_list.xlsx"

# Ordered by prestige priority (1 = highest)
NIFTY_FILES = [
    {
        "path": r"C:\Users\adity\OneDrive\Desktop\MW-NIFTY-50-23-May-2026.csv",
        "index": "Nifty 50",
        "priority": 1,
    },
    {
        "path": r"C:\Users\adity\OneDrive\Desktop\MW-NIFTY-500-23-May-2026.csv",
        "index": "Nifty 500",
        "priority": 2,
    },
    {
        "path": r"C:\Users\adity\OneDrive\Desktop\MW-NIFTY-MIDCAP-150-23-May-2026.csv",
        "index": "Nifty Midcap 150",
        "priority": 3,
    },
    {
        "path": r"C:\Users\adity\OneDrive\Desktop\MW-NIFTY-MIDSMALLCAP-400-23-May-2026.csv",
        "index": "Nifty MidSmallcap 400",
        "priority": 4,
    },
    {
        "path": r"C:\Users\adity\OneDrive\Desktop\MW-NIFTY-MIDCAP-SELECT-23-May-2026.csv",
        "index": "Nifty Midcap Select",
        "priority": 5,
    },
    {
        "path": r"C:\Users\adity\OneDrive\Desktop\MW-NIFTY-MICROCAP-250-23-May-2026.csv",
        "index": "Nifty Microcap 250",
        "priority": 6,
    },
]


# ── helpers ───────────────────────────────────────────────────────────────────

def find_symbol_col(df: pd.DataFrame) -> str:
    for col in df.columns:
        if "SYMBOL" in col.upper():
            return col
    return df.columns[0]


def clean_symbol(raw: str) -> str:
    s = str(raw).strip().upper()
    s = re.sub(r"[^A-Z0-9\-&]", "", s)
    return s


def fetch_company_name(ticker: str) -> tuple[str, str | None]:
    try:
        info = yf.Ticker(ticker).info
        name = info.get("longName") or info.get("shortName")
        return ticker, name
    except Exception:
        return ticker, None


def symbol_to_title(ticker: str) -> str:
    """Fallback: turn 'HDFCBANK.NS' into 'Hdfcbank' as a readable placeholder."""
    return ticker.replace(".NS", "").title()


# ── Step 1 — read CSVs ────────────────────────────────────────────────────────

def read_csv_file(info: dict) -> pd.DataFrame | None:
    path = info["path"]
    label = info["index"]

    if not os.path.exists(path):
        print(f"WARNING: Could not find {os.path.basename(path)}. Skipping and continuing with others.")
        return None

    df = pd.read_csv(path)

    # Strip embedded newlines from column headers
    df.columns = [c.strip().replace("\n", "").strip() for c in df.columns]

    sym_col = find_symbol_col(df)

    # Row 0 is the index banner row (e.g., "NIFTY 50") — skip it
    df = df.iloc[1:].copy()

    raw_syms = df[sym_col].astype(str).str.strip()
    tickers = raw_syms.apply(clean_symbol)
    tickers = tickers + ".NS"

    out = pd.DataFrame({
        "Ticker": tickers,
        "Company": "",
        "Sector": "",
        "Index": label,
        "_priority": info["priority"],
    })

    # Drop rows that are blank or look like summary rows
    out = out[out["Ticker"].str.len() > 3]
    out = out[out["Ticker"] != ".NS"]
    out = out.reset_index(drop=True)

    print(f"[OK] {label}: {len(out)} companies")
    return out


# ── Step 3 — merge & deduplicate ─────────────────────────────────────────────

def deduplicate(df: pd.DataFrame) -> pd.DataFrame:
    # Sort ascending by priority so keep='first' retains the most prestigious index
    df = df.sort_values("_priority")
    df = df.drop_duplicates(subset="Ticker", keep="first")
    return df.reset_index(drop=True)


# ── Step 4 — yfinance enrichment & verification ───────────────────────────────

def enrich_names(tickers: list[str], max_workers: int = 25) -> dict[str, str | None]:
    name_map: dict[str, str | None] = {}
    total = len(tickers)
    done = 0

    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {pool.submit(fetch_company_name, t): t for t in tickers}
        for future in as_completed(futures):
            ticker, name = future.result()
            name_map[ticker] = name
            done += 1
            if done % 100 == 0 or done == total:
                print(f"  Names fetched: {done}/{total}")

    return name_map


def verify_tickers(tickers: list[str], n: int = 5) -> None:
    print("\nVerifying yfinance compatibility...")
    sample = random.sample(tickers, min(n, len(tickers)))
    for ticker in sample:
        try:
            info = yf.Ticker(ticker).info
            name = info.get("longName") or info.get("shortName") or "N/A"
            print(f"  {ticker} -> [OK] Found: {name}")
        except Exception as exc:
            print(f"  {ticker} -> [FAIL] {exc}")


# ── Excel helpers ─────────────────────────────────────────────────────────────

NAVY = PatternFill(fill_type="solid", fgColor="1B3A6B")
WHITE_BOLD = Font(bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center")


def format_sheet(ws) -> None:
    for cell in ws[1]:
        cell.fill = NAVY
        cell.font = WHITE_BOLD
        cell.alignment = CENTER

    ws.freeze_panes = "A2"

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 80)


def add_search_col_to_existing_sheet(ws, ticker_col: int = 1, company_col: int = 2) -> None:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "Search" in headers:
        return
    search_col = ws.max_column + 1
    hdr_cell = ws.cell(1, search_col, "Search")
    hdr_cell.fill = NAVY
    hdr_cell.font = WHITE_BOLD
    hdr_cell.alignment = CENTER
    for row in range(2, ws.max_row + 1):
        company = ws.cell(row, company_col).value or ""
        ticker = ws.cell(row, ticker_col).value or ""
        ws.cell(row, search_col, f"{company} ({ticker})")
    # Autofit the new column
    max_len = max(
        len(str(ws.cell(r, search_col).value or "")) for r in range(1, ws.max_row + 1)
    )
    ws.column_dimensions[get_column_letter(search_col)].width = min(max_len + 4, 80)


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    # ── Step 1 — inspect first file ───────────────────────────────────────────
    first_path = NIFTY_FILES[0]["path"]
    if os.path.exists(first_path):
        _preview = pd.read_csv(first_path)
        _preview.columns = [c.strip().replace("\n", "").strip() for c in _preview.columns]
        print("=== CSV structure (first file preview) ===")
        print(f"Columns: {list(_preview.columns)}")
        print(_preview.head(3).to_string(index=False))
        print("==========================================\n")

    # ── Step 1 — read all files ───────────────────────────────────────────────
    print("Reading NSE CSV files...")
    frames = []
    for info in NIFTY_FILES:
        df = read_csv_file(info)
        if df is not None:
            frames.append(df)

    if not frames:
        print("ERROR: Could not load any CSV files. Please check file paths.")
        sys.exit(1)

    # ── Step 2 — clean tickers ────────────────────────────────────────────────
    print("\nCleaning and formatting tickers...")
    combined = pd.concat(frames, ignore_index=True)
    print("[OK] Added .NS suffix to all tickers")

    # ── Step 3 — deduplicate ──────────────────────────────────────────────────
    print("\nDeduplicating...")
    combined = deduplicate(combined)
    print(f"[OK] {len(combined):,} unique companies after deduplication")

    all_tickers = combined["Ticker"].tolist()

    # ── Step 4 — enrich company names ─────────────────────────────────────────
    print(f"\nFetching company names from yfinance ({len(all_tickers)} tickers, ~2-3 min)...")
    name_map = enrich_names(all_tickers)

    found = sum(1 for v in name_map.values() if v)
    print(f"[OK] Company names resolved: {found}/{len(all_tickers)} from yfinance, "
          f"{len(all_tickers)-found} using symbol fallback")

    combined["Company"] = combined["Ticker"].apply(
        lambda t: name_map.get(t) or symbol_to_title(t)
    )

    # Sort alphabetically by company name
    combined = combined.sort_values("Company", ignore_index=True)
    combined = combined[["Ticker", "Company", "Sector", "Index"]]

    # ── Step 4 — verify 5 random tickers ──────────────────────────────────────
    verify_tickers(all_tickers)

    # ── Step 6 — add Search column ────────────────────────────────────────────
    combined["Search"] = combined["Company"] + " (" + combined["Ticker"] + ")"

    # ── Step 5 — save to Excel ────────────────────────────────────────────────
    print(f"\nSaving to company_list.xlsx...")

    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE)

        # Add Search column to US Companies sheet if present
        if "US Companies" in wb.sheetnames:
            add_search_col_to_existing_sheet(wb["US Companies"], ticker_col=1, company_col=2)

        # Replace India Companies sheet if it already exists
        if "India Companies" in wb.sheetnames:
            del wb["India Companies"]

        ws_india = wb.create_sheet("India Companies")
    else:
        from openpyxl import Workbook
        wb = Workbook()
        ws_india = wb.active
        ws_india.title = "India Companies"

    # Write header
    cols = ["Ticker", "Company", "Sector", "Index", "Search"]
    for c_idx, col_name in enumerate(cols, start=1):
        ws_india.cell(1, c_idx, col_name)

    # Write data rows
    for r_idx, row in enumerate(combined.itertuples(index=False), start=2):
        ws_india.cell(r_idx, 1, row.Ticker)
        ws_india.cell(r_idx, 2, row.Company)
        ws_india.cell(r_idx, 3, row.Sector)
        ws_india.cell(r_idx, 4, row.Index)
        ws_india.cell(r_idx, 5, row.Search)

    format_sheet(ws_india)
    wb.save(OUTPUT_FILE)

    print(f"[OK] India Companies sheet saved: {len(combined):,} companies")
    print(f"\nColumn names in 'India Companies' sheet: Ticker | Company | Sector | Index | Search")
    print(f"Column names in 'US Companies' sheet:    Ticker | Company | Sector | Index | Search")


if __name__ == "__main__":
    main()
