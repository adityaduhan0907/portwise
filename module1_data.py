#!/usr/bin/env python3
"""
module1_data.py  —  Split Estimation (Momentum Returns + Long-Run Risk)

METHODOLOGY
  Expected returns  : 12M-1M momentum window.
                      Monthly prices fetched ~16 months back; the most recent
                      month is skipped to avoid short-term reversal, leaving an
                      11-month holding-period return annualised to 12 months.
  Covariance matrix : 10 years of monthly data for stable long-run risk.
  Display data      : 3 years of daily adjusted closing prices (prices.xlsx).

OUTPUTS
  prices.xlsx         — 3-year daily adjusted closing prices
  returns_stats.xlsx  — 6 sheets:
      "Daily Returns"           daily % returns (3y, display only)
      "Momentum Returns"        monthly % returns in the 11-month window
      "LongRun Monthly Returns" 10-year monthly % returns
      "Stats & Correlation"     per-stock stats + correlation (labeled)
      "Annualised Mu"           pre-computed expected-return vector
      "Annualised Cov"          pre-computed covariance matrix
  (module2 and module3 read "Annualised Mu" and "Annualised Cov" directly)
"""

import json
import math
import os
import pickle
import warnings
from datetime import date, datetime

import numpy as np
import pandas as pd

import fetch_util
from fetch_util import TransientFetchError

warnings.filterwarnings("ignore")

# ── Configuration ──────────────────────────────────────────────────────────────
TRADING_DAYS      = 252
TRADING_MONTHS    = 12
DAILY_YEARS       = 3
LONGRUN_YEARS     = 10
SKIP_RECENT       = 1      # months to skip at near end (short-term reversal)
MOM_WINDOW        = 11     # holding-period months = 12 − 1
MIN_DATA_FRAC     = 0.70
PRICES_MAX_AGE_HRS = 24     # reuse cached raw price history if younger than this
SCRIPT_DIR        = os.path.dirname(os.path.abspath(__file__))

# ── Raw price-history cache (avoids re-fetching yfinance within PRICES_MAX_AGE_HRS) ─
PRICES_CACHE_PATH = os.path.join(SCRIPT_DIR, "prices_cache.pkl")

# ── Factor-model data source ────────────────────────────────────────────────────
DATA_POINTS_PATH = os.path.join(SCRIPT_DIR, "data_points.xlsx")

# ── Sector handoff (here -> robustness_checks Layer-6 concentration check) ───────
SECTORS_PATH     = os.path.join(SCRIPT_DIR, "sectors.json")  # {ticker: sector}

# ── Factor-history handoff (from module 0) + residual handoff (to Part B / Layer 4)
FACTOR_HISTORY_PATH   = os.path.join(SCRIPT_DIR, "factor_history.json")    # module0 -> here
FACTOR_RESIDUALS_PATH = os.path.join(SCRIPT_DIR, "factor_residuals.json")  # here -> Part B
RESIDUAL_START        = "1960-01-01"   # fetch max monthly history from here onward

# Category-beta key -> factor-file column NAME (align betas to factors BY NAME).
# US betas use "MKT"; the matching factor column is named "Mkt-RF".
US_BETA_TO_FACTOR = {"MKT": "Mkt-RF", "SMB": "SMB", "HML": "HML"}
IN_BETA_TO_FACTOR = {"MF": "MF", "SMB": "SMB", "HML": "HML", "WML": "WML"}

# Size split thresholds — native currency, NO FX conversion
US_SIZE_SPLIT    = 6e9       # $6 billion (USD market cap)
INDIA_SIZE_SPLIT = 1.5e12    # ₹1.5 trillion (INR market cap)

# (size, pb_class) -> column letter in data_points.xlsx
#   US 3-factor    (cols B-G): rows 2=MKT 3=SMB 4=HML, premiums row 6
US_BUCKET_COL = {
    ("small", "growth"):  "B",   ("small", "value"):   "C",
    ("big",   "growth"):  "D",   ("big",   "value"):   "E",
    ("small", "neutral"): "F",   ("big",   "neutral"): "G",
}
#   India 4-factor (cols J-O): rows 2=MF 3=SMB 4=HML 5=WML, premiums row 7
INDIA_BUCKET_COL = {
    ("big",   "neutral"): "J",   ("big",   "growth"):  "K",
    ("small", "value"):   "L",   ("small", "neutral"): "M",
    ("small", "growth"):  "N",   ("big",   "value"):   "O",
}
# Neutral-bucket fallback columns used when P/B is unavailable
US_PB_FALLBACK    = {"big": "G", "small": "F"}
INDIA_PB_FALLBACK = {"big": "J", "small": "M"}


def _load_risk_free_rate(fallback=0.043):
    """Load blended risk-free rate from risk_free_rates.json, or fall back to default."""
    json_path = os.path.join(SCRIPT_DIR, "risk_free_rates.json")
    try:
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)
        rate       = float(data["blended_rate"])
        fetched_at = data.get("fetched_at", "unknown")
        print(f"  Risk-free rate : {rate*100:.4f}%  "
              f"(blended, risk_free_rates.json -- fetched {fetched_at})")
        return rate
    except FileNotFoundError:
        print(f"  WARNING: risk_free_rates.json not found -- using fallback {fallback*100:.1f}%.")
        print("           Run module0_riskfree.py first for fresh rates.")
        return fallback
    except Exception as exc:
        print(f"  WARNING: Could not read risk_free_rates.json ({exc}). "
              f"Using fallback {fallback*100:.1f}%.")
        return fallback


RISK_FREE_RATE = _load_risk_free_rate()

def _load_tickers():
    """
    Return the ticker list to process.
    When called from run_all.py, reads the user's tickers from run_config.json.
    Falls back to the hardcoded default when run standalone.
    """
    try:
        cfg_path = os.path.join(SCRIPT_DIR, "run_config.json")
        with open(cfg_path, encoding="utf-8") as f:
            cfg = json.load(f)
        tickers = cfg.get("tickers")
        if tickers:
            print(f"  Tickers from run_config.json: {', '.join(tickers)}")
            return tickers
    except FileNotFoundError:
        pass
    except Exception as exc:
        print(f"  WARNING: Could not read run_config.json for tickers ({exc}).")
    # Default standalone ticker list (edit here for standalone use)
    return [
        "AAPL",
        "MSFT",
        "GOOGL",
        "RELIANCE",    # NSE -- auto-resolved to RELIANCE.NS
        "TCS",         # NSE -- auto-resolved to TCS.NS
        "INFY.NS",     # already has suffix
        "HDFCBANK",    # NSE -- auto-resolved to HDFCBANK.NS
        "WIPRO.BO",    # BSE -- already has suffix
    ]

# ─────────────────────────────────────────────────────────────────────────────
TICKERS = _load_tickers()
# ─────────────────────────────────────────────────────────────────────────────


# ── Download helpers ───────────────────────────────────────────────────────────

def resolve_and_fetch(raw_ticker, start, end, interval="1d", prefer=None):
    """
    Hardened price fetch (retry + backoff + timeout) via fetch_util.

    Tries the ticker as supplied; for a genuinely unsuffixed ticker it falls back
    to .NS then .BO (a suffixed ticker is fetched AS-IS — never the bare US symbol).
    ``prefer`` short-circuits resolution to a symbol already resolved elsewhere
    (e.g. module 0's handoff), so bare symbols are not re-probed within a run.

    Returns (resolved_symbol, series, error_message); error_message is set only
    for GENUINE missing data. A persistent TRANSIENT failure (throttle/timeout/
    connection) RAISES TransientFetchError so it is never silently swallowed.
    """
    symbol, series, err = fetch_util.resolve_and_fetch(
        raw_ticker, start=start, end=end, interval=interval, prefer=prefer,
    )
    if symbol is not None and symbol != raw_ticker and prefer is None:
        print(f"    Auto-resolved '{raw_ticker}' -> '{symbol}'")
    return symbol, series, err


# ── Momentum return (12M-1M) ───────────────────────────────────────────────────

def compute_momentum_return(monthly_prices):
    """
    Standard 12M-1M momentum:
      Skip the SKIP_RECENT most-recent monthly close(s) to avoid reversal.
      Compute the MOM_WINDOW-month holding return, annualise to 12 months.

    Requires >= (SKIP_RECENT + MOM_WINDOW + 2) = 14 monthly price points.
    Returns annualised return (decimal), or float('nan') if insufficient data.

    With SKIP_RECENT=1, MOM_WINDOW=11 and prices sorted ascending:
      end_price   = prices[-3]   (2 months ago)
      start_price = prices[-14]  (13 months ago)
      holding     = 11 months
    """
    prices = monthly_prices.sort_index().dropna()
    needed = SKIP_RECENT + MOM_WINDOW + 2
    if len(prices) < needed:
        return float("nan")

    end_price   = float(prices.iloc[-(SKIP_RECENT + 2)])
    start_price = float(prices.iloc[-(SKIP_RECENT + MOM_WINDOW + 2)])

    if start_price <= 0:
        return float("nan")

    total_return = end_price / start_price - 1
    ann_return   = (1.0 + total_return) ** (12.0 / MOM_WINDOW) - 1
    return ann_return


# ── Per-stock statistics (split estimation) ────────────────────────────────────

def annualised_stats_split(mom_return, longrun_monthly_returns):
    """
    Expected return from momentum window; volatility from long-run monthly data.
    """
    clean   = longrun_monthly_returns.dropna()
    ann_vol = clean.std() * np.sqrt(TRADING_MONTHS)

    if np.isnan(mom_return) or ann_vol == 0:
        sharpe = float("nan")
    else:
        sharpe = (mom_return - RISK_FREE_RATE) / ann_vol

    return {
        "Ann. Return % (Momentum 12M-1M)": (
            round(mom_return * 100, 2) if not np.isnan(mom_return) else float("nan")
        ),
        "Ann. Volatility % (10Y Monthly)": round(ann_vol * 100, 2),
        "Sharpe Ratio":                    (
            round(sharpe, 4) if not np.isnan(sharpe) else float("nan")
        ),
    }


# ── Factor-model expected returns (Fama-French) ─────────────────────────────────

def _to_float(x):
    """Coerce to float; return None for None / NaN / garbage."""
    try:
        if x is None:
            return None
        f = float(x)
        return None if math.isnan(f) else f
    except (TypeError, ValueError):
        return None


def _load_factor_model(path=DATA_POINTS_PATH):
    """
    Read factor betas and premiums from data_points.xlsx in one open.
    Betas are raw; RF + factor premiums are stored as percents (-> /100).
    Returns {"us_betas", "us_prem", "in_betas", "in_prem"}.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    val = lambda coord: float(ws[coord].value)

    # US 3-factor betas: cols B-G, rows 2=MKT 3=SMB 4=HML
    us_betas = {
        col: {"MKT": val(f"{col}2"), "SMB": val(f"{col}3"), "HML": val(f"{col}4")}
        for col in "BCDEFG"
    }
    # US premiums (row 6): B6=RF C6=MKT D6=SMB E6=HML
    us_prem = {
        "RF":  val("B6") / 100.0, "MKT": val("C6") / 100.0,
        "SMB": val("D6") / 100.0, "HML": val("E6") / 100.0,
    }
    # India 4-factor betas: cols J-O, rows 2=MF 3=SMB 4=HML 5=WML
    in_betas = {
        col: {"MF": val(f"{col}2"), "SMB": val(f"{col}3"),
              "HML": val(f"{col}4"), "WML": val(f"{col}5")}
        for col in "JKLMNO"
    }
    # India premiums (row 7): B7=RF C7=MF D7=SMB E7=HML F7=WML
    in_prem = {
        "RF":  val("B7") / 100.0, "MF":  val("C7") / 100.0,
        "SMB": val("D7") / 100.0, "HML": val("E7") / 100.0,
        "WML": val("F7") / 100.0,
    }
    wb.close()
    return {"us_betas": us_betas, "us_prem": us_prem,
            "in_betas": in_betas, "in_prem": in_prem}


def _classify_size(is_india, market_cap):
    """big / small by native-currency market cap (no FX conversion)."""
    split = INDIA_SIZE_SPLIT if is_india else US_SIZE_SPLIT
    if market_cap is None:
        return "small"   # conservative default when market cap is unavailable
    return "big" if market_cap >= split else "small"


def _classify_pb(pb):
    """value (<1.5) / neutral (1.5-4) / growth (>4); None if P/B unavailable."""
    if pb is None or pb <= 0:
        return None
    if pb < 1.5:
        return "value"
    if pb > 4.0:
        return "growth"
    return "neutral"


def _assign_bucket(is_india, size, pb_class):
    """Return (column_letter, human_label). P/B missing -> neutral fallback."""
    if pb_class is None:
        col = (INDIA_PB_FALLBACK if is_india else US_PB_FALLBACK)[size]
        return col, f"{size}+neutral (P/B missing -> col {col})"
    table = INDIA_BUCKET_COL if is_india else US_BUCKET_COL
    return table[(size, pb_class)], f"{size}+{pb_class}"


def _factor_expected_return(is_india, market_cap, pb, model):
    """
    Bucket one stock and return its factor-model expected return plus the
    components needed for display/export.
    """
    size       = _classify_size(is_india, market_cap)
    pb_class   = _classify_pb(pb)
    col, label = _assign_bucket(is_india, size, pb_class)

    if is_india:
        b, p = model["in_betas"][col], model["in_prem"]
        er = (p["RF"] + b["MF"] * p["MF"] + b["SMB"] * p["SMB"]
              + b["HML"] * p["HML"] + b["WML"] * p["WML"])
        beta_mkt, mkt_prem = b["MF"], p["MF"]
        # Betas keyed by the factor-file column NAMES (India sheet order).
        betas_by_factor = {IN_BETA_TO_FACTOR[k]: b[k] for k in ("MF", "SMB", "HML", "WML")}
    else:
        b, p = model["us_betas"][col], model["us_prem"]
        er = (p["RF"] + b["MKT"] * p["MKT"] + b["SMB"] * p["SMB"]
              + b["HML"] * p["HML"])
        beta_mkt, mkt_prem = b["MKT"], p["MKT"]
        # "MKT" beta aligns to the factor named "Mkt-RF".
        betas_by_factor = {US_BETA_TO_FACTOR[k]: b[k] for k in ("MKT", "SMB", "HML")}

    return {
        "market":          "India" if is_india else "USA",
        "size":            size,
        "pb_class":        pb_class if pb_class is not None else "missing",
        "column":          col,
        "bucket":          label,
        "beta_mkt":        beta_mkt,
        "betas_by_factor": betas_by_factor,
        "rf":              p["RF"],
        "mkt_prem":        mkt_prem,
        "expected_return": er,
    }


def _fetch_cap_pb(ticker):
    """
    Fetch (marketCap, priceToBook, sector) from yfinance ticker.info.
    If priceToBook is missing, compute currentPrice / bookValue.
    Sector is captured from the same single info fetch (no extra request) so
    the Layer-6 robustness sector-concentration check has a data source; a
    missing/blank sector is returned as None (callers map it to "Unknown").
    Returns (market_cap_or_None, pb_or_None, sector_or_None).
    """
    # Hardened + memoised .info (retry/backoff on throttle; raises
    # TransientFetchError if the CALL itself keeps failing). A SPARSE dict that
    # merely lacks fields is fine -- the field fallbacks below stay legitimate.
    info = fetch_util.fetch_info(ticker)

    market_cap = _to_float(info.get("marketCap"))

    pb = _to_float(info.get("priceToBook"))
    if pb is None:
        cp = _to_float(info.get("currentPrice"))
        bv = _to_float(info.get("bookValue"))
        if cp is not None and bv not in (None, 0):
            pb = cp / bv

    sector = info.get("sector")
    if isinstance(sector, str):
        sector = sector.strip() or None
    else:
        sector = None
    return market_cap, pb, sector


def calculate_capm_returns(tickers, rf_usa):
    """
    Factor-model expected returns (replaces the old single-beta CAPM):

      US    (3-factor FF) : E[r] = RF + bMKT*MKT + bSMB*SMB + bHML*HML
      India (4-factor FF) : E[r] = RF + bMF*MF + bSMB*SMB + bHML*HML + bWML*WML

    Each stock is bucketed by size (native-currency market cap) and P/B
    (value / neutral / growth); the matching column's betas and the row-6
    (US) / row-7 (India) premiums come from data_points.xlsx.
    Market: ticker ending .NS / .BO -> India, otherwise US.

    Output shape is unchanged so module2 and the Excel export keep working:
      {ticker: {"beta", "market", "rf", "erp", "expected_return"}}
    where "beta" is the market-factor loading and "erp" the market premium.
    (rf_usa is retained for signature compatibility; RF now comes from the file.)
    """
    model = _load_factor_model()
    results = {}
    for ticker in tickers:
        is_india = ticker.upper().endswith((".NS", ".BO"))
        market_cap, pb, sector = _fetch_cap_pb(ticker)
        comp = _factor_expected_return(is_india, market_cap, pb, model)
        results[ticker] = {
            "beta":            comp["beta_mkt"],
            "market":          comp["market"],
            "rf":              comp["rf"],
            "erp":             comp["mkt_prem"],
            "expected_return": comp["expected_return"],
            "column":          comp["column"],
            "bucket":          comp["bucket"],
            "betas_by_factor": comp["betas_by_factor"],
            "sector":          sector if sector else "Unknown",
        }
    return results


# ── Factor covariance + per-asset residuals (Part A) ────────────────────────────

def _load_factor_history(path=FACTOR_HISTORY_PATH):
    """
    Load module 0's cleaned factor history (factor_history.json) and rebuild,
    per market, a monthly factor DataFrame and monthly RF Series indexed by
    month Period. Factor returns and RF are already decimals.

    Returns {market: {"factors": DataFrame, "rf": Series, "names": [...]}}.
    Returns {} if the file is missing/unreadable (caller handles gracefully).
    """
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"  WARNING: {os.path.basename(path)} not found -- run module0 first. "
              "Skipping factor covariance / residuals.")
        return {}
    except Exception as exc:
        print(f"  WARNING: Could not read {os.path.basename(path)} ({exc}).")
        return {}

    out = {}
    for market, md in data.get("markets", {}).items():
        idx = pd.PeriodIndex(pd.to_datetime(md["dates"]), freq="M")
        names = [c for c in md["columns"] if c != "RF"]            # factor names, by name
        factors = pd.DataFrame({c: md["factor_returns"][c] for c in names}, index=idx)
        rf = pd.Series(md.get("monthly_rf", []), index=idx, name="RF")
        out[market] = {"factors": factors, "rf": rf, "names": names}
    return out


def compute_factor_covariance(factor_hist):
    """
    Factor covariance matrix F per market over the FULL available history
    (monthly): US -> 3x3 of [Mkt-RF, SMB, HML]; India -> 4x4 of [MF, SMB, HML, WML].
    Returns {market: DataFrame F (labeled rows/cols)}.
    """
    F = {}
    for market, fh in factor_hist.items():
        fac = fh["factors"].dropna()
        if fac.empty:
            continue
        F[market] = fac.cov()          # monthly factor covariance, labeled
    return F


def compute_asset_residuals(symbol, monthly_returns, market, betas_by_factor, factor_hist):
    """
    Category-beta residuals over the asset's MAX available history.

    For every month the asset has a return that also exists in the factor file:
        residual = (asset_return - RF_that_month) - (betas . factors_that_month)
    RF is the per-month factor-file RF (US sheet col E / India col F), matched by
    month -- never a constant. Betas are the fixed category betas (NOT OLS),
    aligned to factor columns BY NAME.

    Returns (residual_series_indexed_by_Period, n_asset_months).
    """
    n_asset_months = int(monthly_returns.shape[0])
    fh = factor_hist.get(market)
    if fh is None or monthly_returns.empty:
        return pd.Series(dtype=float), n_asset_months

    factors, rf = fh["factors"], fh["rf"]
    common = monthly_returns.index.intersection(factors.index).intersection(rf.index)
    if len(common) == 0:
        return pd.Series(dtype=float), n_asset_months

    excess = monthly_returns.reindex(common) - rf.reindex(common)
    contrib = pd.Series(0.0, index=common)
    for fname, beta in betas_by_factor.items():
        if fname in factors.columns:
            contrib = contrib + beta * factors[fname].reindex(common)

    residual = (excess - contrib).dropna()
    return residual, n_asset_months


def _monthly_returns_from_prices(prices):
    """
    Convert a monthly close Series into monthly returns indexed by month Period.
    Returns an empty Series for missing/empty input. (Math unchanged; split out so
    cached residual-history prices can feed it without a re-fetch.)
    """
    if prices is None or len(prices) == 0:
        return pd.Series(dtype=float)
    rets = prices.sort_index().pct_change().dropna()
    if rets.empty:
        return pd.Series(dtype=float)
    rets.index = pd.PeriodIndex(rets.index, freq="M")
    return rets[~rets.index.duplicated(keep="last")]


def _monthly_returns_max_history(symbol, end, prefer=None):
    """
    Fetch the asset's monthly returns over max history (from RESIDUAL_START),
    indexed by month Period. Returns an empty Series on GENUINE missing data; a
    persistent transient fetch failure raises TransientFetchError (loud).
    """
    _, prices, err = resolve_and_fetch(symbol, RESIDUAL_START, end, "1mo", prefer=prefer)
    if err or prices is None or prices.empty:
        return pd.Series(dtype=float)
    return _monthly_returns_from_prices(prices)


def _load_resolved_map(path=FACTOR_HISTORY_PATH):
    """
    Read module 0's already-resolved symbols (asset_history[ticker]["resolved"])
    from factor_history.json. Lets module 1 fetch the known listing directly via
    ``prefer=`` instead of re-probing bare -> .NS -> .BO a SECOND time this run,
    which both cuts request volume and keeps the two modules on the same listing.
    Returns {REQUESTED_TICKER_UPPER: resolved_symbol}. Empty on any failure.
    """
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        out = {}
        for raw, info in (data.get("asset_history") or {}).items():
            resolved = (info or {}).get("resolved")
            if resolved:
                out[raw.strip().upper()] = resolved
        return out
    except Exception:
        return {}


def _load_use_factor_flag(path=FACTOR_HISTORY_PATH, default=False):
    """
    Read module 0's use_factor_covariance flag from factor_history.json.
    True  -> some portfolio asset has < 36 months -> use factor covariance (B).
    False -> all assets have long history          -> use Ledoit-Wolf (A).
    """
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        return bool(data.get("use_factor_covariance", default))
    except FileNotFoundError:
        print(f"  WARNING: {os.path.basename(path)} not found -- defaulting "
              f"use_factor_covariance={default} (Ledoit-Wolf).")
        return default
    except Exception as exc:
        print(f"  WARNING: Could not read covariance flag ({exc}); "
              f"defaulting to {default}.")
        return default


def _manual_ledoit_wolf(X):
    """
    Ledoit-Wolf (2004) shrinkage toward the scaled identity mu*I -- the same
    target sklearn.covariance.LedoitWolf uses. Pure-numpy fallback for when
    scikit-learn is unavailable. X is (n_samples, n_features), returns
    (covariance ndarray, shrinkage_constant in [0, 1]).
    """
    X = np.asarray(X, dtype=float)
    n, p = X.shape
    Xc = X - X.mean(axis=0)
    S  = (Xc.T @ Xc) / n                      # MLE sample cov (ddof=0)
    mu = np.trace(S) / p
    prior = mu * np.eye(p)

    Y = Xc ** 2
    phi_mat = (Y.T @ Y) / n - S ** 2          # asymptotic variances of S entries
    phi     = phi_mat.sum()
    rho     = np.trace(phi_mat)               # spherical (mu*I) target
    gamma   = np.sum((S - prior) ** 2)        # squared Frobenius misfit to target

    kappa     = (phi - rho) / gamma if gamma > 0 else 0.0
    shrinkage = float(np.clip(kappa / n, 0.0, 1.0))
    sigma     = shrinkage * prior + (1.0 - shrinkage) * S
    return sigma, shrinkage


def _ledoit_wolf_cov(returns_df):
    """
    MONTHLY Ledoit-Wolf shrinkage covariance (Method A). Prefers
    sklearn.covariance.LedoitWolf; falls back to the numpy equivalent.
    Returns (labeled DataFrame, source string).
    """
    cols = list(returns_df.columns)
    X = returns_df[cols].values
    try:
        from sklearn.covariance import LedoitWolf
        lw   = LedoitWolf().fit(X)
        cov  = lw.covariance_
        src  = f"sklearn LedoitWolf (shrinkage={float(lw.shrinkage_):.4f})"
    except Exception:
        cov, shrink = _manual_ledoit_wolf(X)
        src = f"numpy LedoitWolf equiv (shrinkage={shrink:.4f})"
    return pd.DataFrame(cov, index=cols, columns=cols), src


def _factor_model_cov(common, asset_residuals, F_by_market):
    """
    MONTHLY factor-decomposition covariance (Method B): Sigma = B F Bᵀ + D.
      B = category betas (assets x factors), aligned to F BY FACTOR NAME
      F = the market's factor covariance (Part A), D = diag(residual variances)
    Only valid for a SINGLE-market universe. A mixed US+India universe is the
    deferred case -> returns (None, reason) so the caller can guard.
    Returns (labeled DataFrame or None, label/reason).
    """
    markets = {asset_residuals[s]["market"] for s in common if s in asset_residuals}
    if len(markets) != 1:
        return None, f"mixed markets {sorted(markets)} -- deferred"
    market = next(iter(markets))

    F_df = F_by_market.get(market)
    if F_df is None or F_df.empty:
        return None, f"no factor covariance available for {market}"

    fnames = list(F_df.columns)               # factor order defines B's columns
    Fm = F_df.values
    B  = np.array([[float(asset_residuals[s]["betas"].get(fn, 0.0)) for fn in fnames]
                   for s in common])
    dvar = []
    for s in common:
        resid = asset_residuals[s]["residuals"]
        dvar.append(float(resid.var(ddof=1)) if resid.shape[0] > 1 else 0.0)

    sigma = B @ Fm @ B.T + np.diag(dvar)      # monthly asset covariance
    df = pd.DataFrame(sigma, index=common, columns=common)
    return df, f"{market}: {len(fnames)} factors x {len(common)} assets"


def build_asset_covariance(common, cov_clean, returns_df, cov_cols,
                           use_factor_cov, asset_residuals, F_by_market):
    """
    Build the asset covariance and ANNUALISE it (x TRADING_MONTHS) so the output
    contract is identical to the previous 10-year sample covariance that module 2
    reads from the 'Annualised Cov' sheet (same shape, labels, annualised units).

    Method A (use_factor_cov False, DEFAULT): Ledoit-Wolf on 10y monthly returns.
    Method B (use_factor_cov True): factor decomposition B F Bᵀ + D (single market).
                                    Mixed US+India is DEFERRED -> guarded stopgap.
    Returns (annualised covariance DataFrame over `common`, method label).
    """
    # Degenerate: no monthly overlap -> keep the original daily-sample fallback.
    if cov_clean.empty or len(cov_clean) < 2:
        cov = returns_df[cov_cols].dropna().cov() * TRADING_DAYS
        return cov.loc[common, common], "fallback: daily sample cov (no monthly overlap)"

    if use_factor_cov:
        monthly_df, info = _factor_model_cov(common, asset_residuals, F_by_market)
        if monthly_df is not None:
            return (monthly_df.loc[common, common] * TRADING_MONTHS,
                    f"Method B -- factor B F B^T + D  [{info}]")
        # Mixed-market factor covariance (cross-market F) is intentionally deferred.
        print("\n  NOTE: use_factor_covariance=True but the portfolio mixes markets.")
        print(f"        {info}. Cross-market factor covariance is DEFERRED (not solved).")
        print("        Guard: falling back to Ledoit-Wolf (Method A) as a labeled stopgap.")
        monthly_df, src = _ledoit_wolf_cov(cov_clean[common])
        return (monthly_df.loc[common, common] * TRADING_MONTHS,
                f"Method B DEFERRED (mixed) -> Method A stopgap  [{src}]")

    monthly_df, src = _ledoit_wolf_cov(cov_clean[common])
    return (monthly_df.loc[common, common] * TRADING_MONTHS,
            f"Method A -- Ledoit-Wolf  [{src}]")


def write_residuals_handoff(F_by_market, asset_residuals, generated_at,
                            path=FACTOR_RESIDUALS_PATH):
    """
    Persist (same write-a-JSON handoff pattern as module 0):
      - factor covariance F per market (labeled)
      - per asset: category betas, residual VARIANCE (diagonal D for Part B),
        and the FULL residual SHOCK vector + its months (for Layer 4 simulation;
        the simulation re-adds the month's RF during reconstruction).
    """
    cov_payload = {}
    for market, F in F_by_market.items():
        cov_payload[market] = {
            "factors":  list(F.columns),
            "matrix":   [[round(float(x), 12) for x in row] for row in F.values],
            "n_months": int(F.attrs.get("n_months", 0)) or None,
        }

    assets_payload = {}
    for sym, info in asset_residuals.items():
        resid = info["residuals"]
        assets_payload[sym] = {
            "market":            info["market"],
            "bucket":            info["bucket"],
            "column":            info["column"],
            "betas":             {k: round(float(v), 8) for k, v in info["betas"].items()},
            "months_history":    info["months_history"],
            "residual_count":    int(resid.shape[0]),
            "residual_variance": (round(float(resid.var(ddof=1)), 12)
                                  if resid.shape[0] > 1 else None),
            "residual_months":   [str(p) for p in resid.index],
            "residual_shocks":   [round(float(x), 10) for x in resid.values],
        }

    payload = {
        "generated_at":     generated_at,
        "factor_covariance": cov_payload,
        "assets":           assets_payload,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    return path


# ── Raw price-history cache ─────────────────────────────────────────────────────
#   Caches the three raw per-symbol price dicts (daily / momentum-monthly /
#   long-run-monthly) that EVERYTHING module 1 derives from -- the 3y daily
#   prices.xlsx, the momentum window, and the 10y monthly covariance returns -- plus
#   a timestamp and the requested-ticker set so freshness (< PRICES_MAX_AGE_HRS) and
#   coverage (cache covers all requested tickers) can be checked. Analysis math is
#   unchanged; this only avoids the network fetch when a fresh cache already covers
#   the requested tickers.

def _load_price_cache(path=PRICES_CACHE_PATH):
    """Load the price cache (or None if absent/unreadable)."""
    if not os.path.exists(path):
        return None
    try:
        with open(path, "rb") as f:
            return pickle.load(f)
    except Exception as exc:
        print(f"  WARNING: could not read price cache ({exc}); will download.")
        return None


def _cache_age_hours(cache):
    """Hours since the cache was downloaded (inf if unknown)."""
    try:
        dt = datetime.fromisoformat(cache["downloaded_at"])
        return (datetime.now() - dt).total_seconds() / 3600.0
    except Exception:
        return float("inf")


def _cache_usable(cache, tickers):
    """
    True if the cache is fresh (< PRICES_MAX_AGE_HRS) AND covers every requested
    ticker (the cached requested set is a superset of the current request).
    """
    if not cache:
        return False, "no cache"
    age = _cache_age_hours(cache)
    if age > PRICES_MAX_AGE_HRS:
        return False, f"stale ({age:.1f}h > {PRICES_MAX_AGE_HRS}h)"
    cached_req = set(cache.get("requested_tickers", []))
    requested  = {t.strip().upper() for t in tickers}
    if not requested.issubset(cached_req):
        missing = sorted(requested - cached_req)
        return False, f"missing tickers {missing}"
    return True, f"{age:.1f}h"


def _slice_cache(cache, tickers):
    """
    Reconstruct (daily_prices, mom_prices, cov_prices, resid_prices, failed)
    limited to the requested tickers from a usable cache. Keyed by resolved symbol,
    as the fetch loop would have produced. ``resid_prices`` is the max-history
    monthly close series for the residual computation -- absent in pre-upgrade
    caches, in which case it comes back all-None and main() re-fetches it once.
    """
    requested = [t.strip().upper() for t in tickers]
    t2s = cache.get("ticker_to_symbol", {})
    symbols = [t2s[t] for t in requested if t in t2s]

    daily = {s: cache["daily_prices"][s] for s in symbols if s in cache["daily_prices"]}
    mom   = {s: cache["mom_prices"].get(s) for s in symbols}
    cov   = {s: cache["cov_prices"].get(s) for s in symbols}
    resid = {s: cache.get("resid_prices", {}).get(s) for s in symbols}
    failed = {t: cache["failed"][t] for t in requested if t in cache.get("failed", {})}
    return daily, mom, cov, resid, failed


def _save_price_cache(daily_prices, mom_prices, cov_prices, resid_prices, failed,
                      requested_tickers, ticker_to_symbol, path=PRICES_CACHE_PATH):
    """Persist the raw price dicts + timestamp + ticker set for later reuse."""
    payload = {
        "downloaded_at":     datetime.now().isoformat(timespec="seconds"),
        "requested_tickers": [t.strip().upper() for t in requested_tickers],
        "ticker_to_symbol":  dict(ticker_to_symbol),
        "symbols":           list(daily_prices.keys()),
        "daily_prices":      daily_prices,
        "mom_prices":        mom_prices,
        "cov_prices":        cov_prices,
        "resid_prices":      resid_prices,
        "failed":            failed,
    }
    try:
        with open(path, "wb") as f:
            pickle.dump(payload, f)
    except Exception as exc:
        print(f"  WARNING: could not write price cache ({exc}).")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    today = date.today()

    # Date windows
    daily_start = (pd.Timestamp(today)
                   - pd.DateOffset(years=DAILY_YEARS, days=10)).date()
    mom_start   = (pd.Timestamp(today)
                   - pd.DateOffset(months=SKIP_RECENT + MOM_WINDOW + 4)).date()
    cov_start   = (pd.Timestamp(today)
                   - pd.DateOffset(years=LONGRUN_YEARS, months=3)).date()

    min_obs_daily = int(TRADING_DAYS * DAILY_YEARS * MIN_DATA_FRAC)

    W = 70
    print(f"\n{'='*W}")
    print("  Module 1 -- Split Estimation: Momentum Returns + Long-Run Risk")
    print(f"  Daily data   : {daily_start} -> {today}  ({DAILY_YEARS}y)")
    print(f"  Momentum     : {SKIP_RECENT + MOM_WINDOW + 1}M lookback, skip {SKIP_RECENT}M,")
    print(f"                 {MOM_WINDOW}-month holding period, annualised (monthly data)")
    print(f"  Covariance   : {LONGRUN_YEARS}-year monthly data")
    print(f"  Risk-free r  : {RISK_FREE_RATE*100:.4f}%")
    print(f"{'='*W}\n")

    print("  NOTE: Expected returns use the momentum anomaly (12M-1M window).")
    print("        The most recent month is excluded to avoid short-term reversal.")
    print("  NOTE: Risk estimates use 10 years of monthly data for stability.\n")

    # ── 1. Fetch all data (reuse a fresh cache if it covers the tickers) ───────
    daily_prices  = {}   # symbol -> pd.Series (daily)
    mom_prices    = {}   # symbol -> pd.Series (monthly closes, momentum window)
    cov_prices    = {}   # symbol -> pd.Series (monthly closes, long-run)
    resid_prices  = {}   # symbol -> pd.Series (monthly closes, max history; residuals)
    failed        = {}   # original ticker -> error message

    # Symbols module 0 already resolved -> fetch that listing directly (no re-probe).
    resolved_map = _load_resolved_map()

    cache = _load_price_cache()
    usable, reason = _cache_usable(cache, TICKERS)
    if usable:
        print(f"  Using cached prices (downloaded {reason} ago) -- skipping yfinance download.")
        daily_prices, mom_prices, cov_prices, resid_prices, failed = _slice_cache(cache, TICKERS)
    else:
        print(f"  Downloading fresh prices...  (cache: {reason})")
        ticker_to_symbol = {}   # requested ticker -> resolved symbol
        for raw in TICKERS:
            ticker = raw.strip().upper()
            print(f"  [{ticker}]")
            prefer = resolved_map.get(ticker)   # module 0's resolved listing, if any

            # Daily (3 years) -- determines the resolved symbol
            symbol, daily_s, err = resolve_and_fetch(ticker, daily_start, today, "1d",
                                                     prefer=prefer)
            if err:
                print(f"    SKIP: {err}")
                failed[ticker] = err
                continue
            if len(daily_s) < min_obs_daily:
                err = (f"only {len(daily_s)} daily obs, need >= {min_obs_daily} "
                       f"({MIN_DATA_FRAC:.0%} of {DAILY_YEARS}y)")
                print(f"    SKIP: {err}")
                failed[ticker] = err
                continue

            ticker_to_symbol[ticker] = symbol
            daily_prices[symbol] = daily_s
            print(f"    Daily      : {len(daily_s):>4} obs  OK")

            # Momentum monthly (reuse resolved symbol -> single, suffixed candidate)
            _, mom_s, err_m = resolve_and_fetch(symbol, mom_start, today, "1mo")
            if err_m or mom_s is None:
                print(f"    Momentum   : WARNING -- {err_m or 'no data'}")
                mom_prices[symbol] = None
            else:
                needed = SKIP_RECENT + MOM_WINDOW + 2
                if len(mom_s) < needed:
                    print(f"    Momentum   : WARNING -- only {len(mom_s)} monthly prices "
                          f"(need {needed}); momentum return will use fallback")
                else:
                    print(f"    Momentum   : {len(mom_s):>4} monthly prices  OK")
                mom_prices[symbol] = mom_s

            # Long-run monthly (reuse resolved symbol)
            _, cov_s, err_c = resolve_and_fetch(symbol, cov_start, today, "1mo")
            if err_c or cov_s is None:
                print(f"    Long-run   : WARNING -- {err_c or 'no data'}; will use available data")
                cov_prices[symbol] = None
            else:
                min_cov = int(LONGRUN_YEARS * TRADING_MONTHS * MIN_DATA_FRAC)
                if len(cov_s) < min_cov:
                    print(f"  WARNING: {symbol} only has {len(cov_s)} months of data. "
                          f"Using full available history.")
                else:
                    print(f"    Long-run   : {len(cov_s):>4} monthly prices  OK")
                cov_prices[symbol] = cov_s

            # Residual max-history monthly (reuse resolved symbol) -- fetched here
            # (instead of an uncached re-pull in section 6b) so a cached run is
            # genuinely network-light. None on genuine missing; transient raises.
            _, resid_s, err_r = resolve_and_fetch(symbol, RESIDUAL_START, today, "1mo")
            if err_r or resid_s is None:
                print(f"    Residual   : WARNING -- {err_r or 'no data'}; residuals will be empty")
                resid_prices[symbol] = None
            else:
                print(f"    Residual   : {len(resid_s):>4} monthly prices  OK")
                resid_prices[symbol] = resid_s

        # Persist the raw history so the next run within PRICES_MAX_AGE_HRS reuses it.
        if daily_prices:
            _save_price_cache(daily_prices, mom_prices, cov_prices, resid_prices, failed,
                              TICKERS, ticker_to_symbol)

    if not daily_prices:
        print("\n  No tickers processed successfully. Exiting.\n")
        return

    symbols = list(daily_prices.keys())

    # ── 2. Build DataFrames ────────────────────────────────────────────────────
    prices_df  = pd.DataFrame(daily_prices).sort_index()
    returns_df = prices_df.pct_change().iloc[1:]

    # Momentum monthly returns (only the window rows, for the Excel sheet)
    mom_dfs = {s: v for s, v in mom_prices.items() if v is not None}
    if mom_dfs:
        mom_prices_df  = pd.DataFrame(mom_dfs).sort_index()
        mom_returns_df = mom_prices_df.pct_change().iloc[1:]
        # Keep only the rows within the momentum window (last MOM_WINDOW + SKIP_RECENT + 1)
        keep = MOM_WINDOW + SKIP_RECENT + 2
        mom_window_df  = mom_returns_df.tail(keep)
    else:
        mom_prices_df  = pd.DataFrame()
        mom_window_df  = pd.DataFrame()

    # Long-run monthly returns
    cov_dfs = {s: v for s, v in cov_prices.items() if v is not None}
    if cov_dfs:
        cov_prices_df = pd.DataFrame(cov_dfs).sort_index()
        cov_rets_df   = cov_prices_df.pct_change().iloc[1:]
    else:
        print("  WARNING: No long-run monthly data -- falling back to daily returns for covariance.")
        cov_rets_df = returns_df

    # ── 3. Compute mu (momentum) ───────────────────────────────────────────────
    mu_dict = {}
    for symbol in symbols:
        m_prices = mom_prices.get(symbol)
        if m_prices is not None:
            r = compute_momentum_return(m_prices)
        else:
            r = float("nan")

        if np.isnan(r):
            # Fallback: annualised mean of available daily returns
            if symbol in returns_df.columns:
                r = returns_df[symbol].dropna().mean() * TRADING_DAYS
                print(f"  [{symbol}] Momentum fallback: using daily-based annualised return.")
            else:
                r = 0.0
        mu_dict[symbol] = r

    # ── 4. Asset return series for covariance (10y monthly) ────────────────────
    #        (the covariance MATRIX itself is built later, after the factor
    #         structures, so Method B can use them -- see "Covariance" below.)
    cov_cols  = [s for s in symbols if s in cov_rets_df.columns]
    cov_clean = cov_rets_df[cov_cols].dropna()
    common    = list(cov_cols)

    # ── 5. Per-stock stats and correlation ────────────────────────────────────
    stats_rows = {}
    for s in common:
        lr_rets = cov_clean[s] if s in cov_clean.columns else pd.Series(dtype=float)
        stats_rows[s] = annualised_stats_split(mu_dict[s], lr_rets)

    stats_df = pd.DataFrame(stats_rows).T
    stats_df.index.name = "Ticker"

    corr_df = cov_clean.corr() if not cov_clean.empty else returns_df[cov_cols].corr()

    # ── 6. Factor-model expected returns ───────────────────────────────────────
    print(f"\n{'='*W}")
    print( "  FACTOR-MODEL EXPECTED RETURNS  (US: 3-factor FF | India: 4-factor FF+WML)")
    print(f"{'='*W}")
    capm_results = calculate_capm_returns(common, RISK_FREE_RATE)
    for s, v in capm_results.items():
        print(f"    {s:<22}  [{v['market']:5}]  beta={v['beta']:.3f}  "
              f"Rf={v['rf']*100:.2f}%  ERP={v['erp']*100:.2f}%  "
              f"E[r]={v['expected_return']*100:>7.2f}%")

    # Persist ticker -> sector for the Layer-6 robustness concentration check.
    # Captured from the same ticker.info fetch above (marketCap / priceToBook);
    # unresolved sectors are stored as "Unknown".
    sectors_map = {t: v.get("sector", "Unknown") for t, v in capm_results.items()}
    try:
        with open(SECTORS_PATH, "w", encoding="utf-8") as f:
            json.dump(sectors_map, f, indent=2)
        n_unknown = sum(1 for s in sectors_map.values() if s == "Unknown")
        print(f"  Sectors saved -> sectors.json  ({len(sectors_map)} tickers, "
              f"{n_unknown} Unknown)")
    except Exception as exc:
        print(f"  WARNING: could not write sectors.json ({exc}).")

    capm_df = pd.DataFrame([
        {
            "Ticker":               t,
            "Market":               v["market"],
            "Beta":                 round(v["beta"], 4),
            "CAPM_Expected_Return": round(v["expected_return"], 6),
            "Rf":                   round(v["rf"], 6),
            "ERP":                  round(v["erp"], 6),
        }
        for t, v in capm_results.items()
    ])

    # ── 6b. Factor covariance F + category-beta residuals (Part A) ─────────────
    #        Computed here so the covariance build below (6c) can consume them.
    print(f"\n{'='*W}")
    print( "  FACTOR COVARIANCE (F) + CATEGORY-BETA RESIDUALS  (Part A)")
    print(f"{'='*W}")

    factor_hist = _load_factor_history()
    F_by_market = compute_factor_covariance(factor_hist)
    for market, F in F_by_market.items():
        n_months = len(factor_hist[market]["factors"].dropna())
        F.attrs["n_months"] = n_months
        print(f"\n  Factor covariance F -- {market}  (monthly, {n_months} months, "
              f"{F.shape[0]}x{F.shape[1]})")
        print(F.round(8).to_string())

    asset_residuals = {}
    print(f"\n  Per-asset residuals (max history, betas by factor name):")
    for s in common:
        cr = capm_results.get(s)
        if cr is None:
            continue
        market    = cr["market"]
        betas     = cr["betas_by_factor"]
        # Prefer the residual prices fetched/cached in section 1; only re-fetch if
        # absent (genuine miss, or a pre-upgrade cache without resid_prices).
        cached_resid = resid_prices.get(s)
        if cached_resid is not None:
            m_rets = _monthly_returns_from_prices(cached_resid)
        else:
            m_rets = _monthly_returns_max_history(s, today)
        residuals, n_hist = compute_asset_residuals(s, m_rets, market, betas, factor_hist)
        asset_residuals[s] = {
            "market":         market,
            "bucket":         cr["bucket"],
            "column":         cr["column"],
            "betas":          betas,
            "months_history": n_hist,
            "residuals":      residuals,
        }
        beta_str = ", ".join(f"{k}={v:+.3f}" for k, v in betas.items())
        rvar = residuals.var(ddof=1) if residuals.shape[0] > 1 else float("nan")
        print(f"    {s:<22} [{market:5}]  {cr['bucket']:<22} col {cr['column']}")
        print(f"        betas : {beta_str}")
        print(f"        history={n_hist:>4}mo  residuals={residuals.shape[0]:>4}  "
              f"resid_var={rvar:.6e}")

    try:
        rp = write_residuals_handoff(F_by_market, asset_residuals, today.isoformat())
        print(f"\n  Saved residual handoff -> {rp}")
    except Exception as exc:
        print(f"\n  ERROR: Could not write residual handoff: {exc}")

    # ── 6c. Covariance (Method A Ledoit-Wolf default / Method B factor fallback) ─
    #        Switches on module 0's use_factor_covariance flag. Output is the
    #        annualised, ticker-labeled matrix module 2 reads from 'Annualised Cov'
    #        -- same shape/labels/units as before (the optimizer is unaffected).
    use_factor_cov = _load_use_factor_flag()
    cov_matrix, cov_method = build_asset_covariance(
        common, cov_clean, returns_df, cov_cols,
        use_factor_cov, asset_residuals, F_by_market,
    )

    print(f"\n{'='*W}")
    print("  COVARIANCE MATRIX  (Layer 2 / Part B)")
    print(f"{'='*W}")
    print(f"  use_factor_covariance (module0 flag) : {use_factor_cov}")
    print(f"  Method                               : {cov_method}")
    print(f"  Shape / labels                       : {cov_matrix.shape}  "
          f"[{', '.join(map(str, cov_matrix.columns))}]")
    print(f"  Units                                : annualised (monthly x {TRADING_MONTHS})")
    diag = {s: round(float(cov_matrix.loc[s, s]), 6) for s in common}
    print(f"  Diagonal (annualised variances)      : {diag}")
    if len(common) >= 2:
        a, b = common[0], common[1]
        print(f"  Sample off-diagonals                 : "
              f"cov[{a},{b}]={cov_matrix.loc[a, b]:+.6f}", end="")
        if len(common) >= 3:
            c = common[2]
            print(f" | cov[{a},{c}]={cov_matrix.loc[a, c]:+.6f} "
                  f"| cov[{b},{c}]={cov_matrix.loc[b, c]:+.6f}", end="")
        print()

    # ── 7. Prepare Annualised Mu and Cov for export ───────────────────────────
    mu_series  = pd.Series({s: mu_dict[s] for s in common}, name="Annualised_Expected_Return")
    mu_series.index.name = "Ticker"
    cov_export = cov_matrix.loc[common, common]

    # ── 7. Console output ──────────────────────────────────────────────────────
    print(f"\n{'='*W}")
    print("  INDIVIDUAL STOCK STATISTICS")
    print("  Return source : momentum window (12M-1M, monthly data)")
    print("  Risk source   : long-run monthly standard deviation (10y)")
    print(f"{'='*W}")
    print(stats_df.to_string())

    print(f"\n{'='*W}")
    print("  ANNUALISED EXPECTED RETURNS  (momentum 12M-1M)")
    print(f"{'='*W}")
    for s, r in mu_dict.items():
        print(f"    {s:<22}  {r*100:>8.2f}%")

    print(f"\n{'='*W}")
    print("  CORRELATION MATRIX  (source: long-run monthly returns)")
    print(f"{'='*W}")
    print(corr_df.round(4).to_string())

    print(f"\n{'='*W}")
    print("  *** MOMENTUM RISK WARNING ***")
    print("  Momentum-based expected returns carry REVERSAL RISK during market")
    print("  regime changes.  High recent performers may underperform sharply")
    print("  when trends reverse.  RERUN THIS MODEL AT LEAST MONTHLY to stay")
    print("  current.  Do not use stale estimates for live portfolio decisions.")
    print(f"{'='*W}\n")

    # ── 8. Excel export ────────────────────────────────────────────────────────
    prices_path  = os.path.join(SCRIPT_DIR, "prices.xlsx")
    returns_path = os.path.join(SCRIPT_DIR, "returns_stats.xlsx")

    prices_df.to_excel(prices_path, index_label="Date")

    note_rows = [
        "METHODOLOGY NOTES",
        f"Expected Return : Momentum window (12M-1M, monthly data; most recent "
        f"{SKIP_RECENT} month skipped for reversal)",
        f"Risk (Volatility): {LONGRUN_YEARS}-year monthly data, annualised (x sqrt(12))",
        f"Covariance method: {cov_method}  (annualised x{TRADING_MONTHS}; switches on module0 use_factor_covariance)",
        f"Risk-free rate   : {RISK_FREE_RATE*100:.4f}% (from risk_free_rates.json)",
        "Module2 / Module3: read 'Annualised Mu' and 'Annualised Cov' sheets directly.",
    ]
    note_df = pd.DataFrame({"Note": note_rows})

    n_stats = len(stats_df)
    gap     = 2
    with pd.ExcelWriter(returns_path, engine="openpyxl") as writer:

        # Sheet 1: Daily Returns
        returns_df.to_excel(writer, sheet_name="Daily Returns", index_label="Date")

        # Sheet 2: Momentum Returns (window rows only)
        if not mom_window_df.empty:
            mom_window_df.to_excel(
                writer, sheet_name="Momentum Returns", index_label="Date"
            )

        # Sheet 3: LongRun Monthly Returns
        if not cov_clean.empty:
            cov_clean.to_excel(
                writer, sheet_name="LongRun Monthly Returns", index_label="Date"
            )

        # Sheet 4: Stats & Correlation  (with methodology header)
        sname = "Stats & Correlation"
        note_df.to_excel(writer, sheet_name=sname, startrow=0, index=False)

        stats_start   = len(note_df) + 2
        corr_lbl_row  = stats_start + n_stats + gap       # 0-based
        corr_data_row = corr_lbl_row + 1                  # 0-based

        stats_df.to_excel(writer, sheet_name=sname, startrow=stats_start)
        ws = writer.sheets[sname]
        ws.cell(row=corr_lbl_row + 1, column=1,
                value="Correlation Matrix  (long-run monthly returns)")
        corr_df.round(4).to_excel(
            writer, sheet_name=sname, startrow=corr_data_row
        )

        # Sheet 5: Annualised Mu  (module2 / module3 source)
        mu_export = mu_series.reset_index()
        mu_export.columns = ["Ticker", "Annualised_Expected_Return"]
        mu_export["Estimation"] = f"Momentum {SKIP_RECENT + MOM_WINDOW + 1}M-{SKIP_RECENT}M (monthly)"
        mu_export.to_excel(writer, sheet_name="Annualised Mu", index=False)

        # Sheet 6: Annualised Cov  (module2 / module3 source)
        cov_export.to_excel(writer, sheet_name="Annualised Cov")

        # Sheet 7: CAPM Returns  (app.py reads this for Table 3B)
        capm_df.to_excel(writer, sheet_name="CAPM Returns", index=False)

    print(f"  Exported -> {prices_path}")
    print(f"  Exported -> {returns_path}")
    print( "    Sheets : Daily Returns | Momentum Returns | LongRun Monthly Returns")
    print( "             Stats & Correlation | Annualised Mu | Annualised Cov | CAPM Returns")

    # ── 9. Summary ─────────────────────────────────────────────────────────────
    print(f"\n{'='*W}")
    print("  SUMMARY")
    print(f"{'='*W}")
    print(f"  Successfully processed : {len(daily_prices)}")
    print(f"  Failed / skipped       : {len(failed)}")
    if failed:
        print()
        for t, reason in failed.items():
            print(f"    {t}: {reason}")
    print(f"{'='*W}\n")


if __name__ == "__main__":
    import sys
    try:
        main()
    except TransientFetchError as exc:
        # Loud, named transient failure -> hand the ticker/call to run_all (parent)
        # and exit non-zero so the pipeline STOPS instead of producing degraded data.
        fetch_util.write_fetch_error(exc.ticker, exc.what, exc.detail, module="module1")
        print(f"\n  ERROR: {exc}\n")
        sys.exit(1)
