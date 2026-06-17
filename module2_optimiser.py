#!/usr/bin/env python3
"""
module2_optimiser.py  —  Layer 3 Core Optimizer

Reads returns_stats.xlsx produced by module1_data.py, then builds three goals:

  1. Minimum Variance (GMV) — minimise wᵀ·Σ·w on module 1's covariance
     (Ledoit-Wolf / factor-decomposition "Annualised Cov"). Constraints:
     w >= 0, sum to 1, NO concentration caps (the naturally diversified
     baseline). Ignores expected returns.

  2. Maximum Risk-Adjusted Return — maximise the Sharpe ratio
     (wᵀ·mu - rf) / sqrt(wᵀ·Σ·w) using the FACTOR expected returns
     (FF3-US / Carhart4-India, "CAPM Returns" sheet), the same Σ, and
     rf = blended_rate. Constraints: w >= 0, sum to 1, PLUS adaptive weight
     caps (3 stocks 60% / 4-6 35% / 7-15 20%) to prevent concentration.

  3. Tail-Risk Minimization — handled by Layer 6 (CVaR), which needs the
     Layer 4 simulation. Neither exists yet, so this is a clean STUB.

Momentum mu (module 1) is still loaded/displayed but is no longer used by any
optimiser goal.

Exports goals 1-2 to optimised_portfolios.xlsx (same weight/stats layout as
before, new goal sheet names).
"""

import json
import os
import sys
import warnings

import numpy as np
import pandas as pd
from scipy.optimize import minimize

import tail_risk_optimizer as tro   # Layer 6 — fills Goal 3 (CVaR)

warnings.filterwarnings("ignore")

# ── Configuration ──────────────────────────────────────────────────────────────
TRADING_DAYS = 252
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))


def get_adaptive_bounds(n):
    """Return (min_weight, max_weight) based on number of stocks."""
    if n == 3:
        return 0.05, 0.60
    elif n <= 6:
        return 0.03, 0.35
    else:
        return 0.02, 0.20


def _load_risk_free_rate(fallback=0.043):
    """Load blended risk-free rate from risk_free_rates.json, or fall back to default."""
    json_path = os.path.join(SCRIPT_DIR, "risk_free_rates.json")
    try:
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)
        rate       = float(data["blended_rate"])
        fetched_at = data.get("fetched_at", "unknown")
        print(f"  Risk-free rate : {rate*100:.4f}%  "
              f"(blended, risk_free_rates.json — fetched {fetched_at})")
        return rate
    except FileNotFoundError:
        print(f"  WARNING: risk_free_rates.json not found — using fallback {fallback*100:.1f}%.")
        print("           Run module0_riskfree.py first for fresh rates.")
        return fallback
    except Exception as exc:
        print(f"  WARNING: Could not read risk_free_rates.json ({exc}). "
              f"Using fallback {fallback*100:.1f}%.")
        return fallback


RISK_FREE_RATE = _load_risk_free_rate()


def load_capm_mu(returns_path, tickers, fallback_mu):
    """
    Load the FACTOR-model expected returns (FF3-US / Carhart4-India) from the
    "CAPM Returns" sheet, aligned to `tickers`. Any ticker missing from the sheet
    falls back to fallback_mu[i]. Shared by main() and the Layer 7 resampler.
    """
    capm_df  = pd.read_excel(returns_path, sheet_name="CAPM Returns")
    capm_map = dict(zip(capm_df["Ticker"], capm_df["CAPM_Expected_Return"]))
    return np.array([capm_map.get(t, fallback_mu[i]) for i, t in enumerate(tickers)])


# ── Helper: portfolio statistics ───────────────────────────────────────────────

def portfolio_stats(weights, mu, cov):
    """Return (annualised return, annualised volatility, Sharpe ratio)."""
    ret = float(weights @ mu)
    vol = float(np.sqrt(weights @ cov @ weights))
    sharpe = (ret - RISK_FREE_RATE) / vol if vol > 0 else float("nan")
    return ret, vol, sharpe


# ── Objective functions (scipy minimises, so we negate to maximise) ────────────

def neg_sharpe(weights, mu, cov):
    _, vol, sharpe = portfolio_stats(weights, mu, cov)
    return -sharpe if vol > 0 else 1e9


def portfolio_vol(weights, mu, cov):
    return float(np.sqrt(weights @ cov @ weights))


# ── Core optimiser ─────────────────────────────────────────────────────────────

def optimise(objective, mu, cov, label, w_min, w_max):
    """
    Run SLSQP minimisation with explicit per-asset bounds.
      w_min/w_max : lower / upper weight bound applied to every asset.
        - GMV (goal 1)            -> (0.0, 1.0)  (no concentration caps)
        - Max Risk-Adjusted (g.2) -> (0.0, adaptive cap)  (caps, no min floor)
    Returns (weights_array, success_bool, message_str).
    """
    n = len(mu)
    bounds      = [(w_min, w_max)] * n
    constraints = [{"type": "eq", "fun": lambda w: np.sum(w) - 1.0}]

    # Try several random starting points; keep best result
    best_result = None
    rng = np.random.default_rng(seed=42)

    for attempt in range(50):
        # Random Dirichlet start, then clip to bounds and renormalise
        w0 = rng.dirichlet(np.ones(n))
        w0 = np.clip(w0, w_min, w_max)
        w0 /= w0.sum()

        res = minimize(
            objective,
            w0,
            args=(mu, cov),
            method="SLSQP",
            bounds=bounds,
            constraints=constraints,
            options={"ftol": 1e-12, "maxiter": 2000},
        )

        if res.success or res.status == 0:
            if best_result is None or res.fun < best_result.fun:
                best_result = res

    if best_result is None or not (best_result.success or best_result.status == 0):
        return None, False, f"Optimisation did not converge for '{label}'"

    weights = np.clip(best_result.x, w_min, w_max)
    weights /= weights.sum()   # re-normalise after clipping
    return weights, True, "OK"


# ── Display & export helpers ───────────────────────────────────────────────────

def build_result_df(tickers, weights, mu, cov):
    """Return a DataFrame summarising portfolio weights and aggregate stats."""
    ret, vol, sharpe = portfolio_stats(weights, mu, cov)

    rows = []
    for ticker, w in zip(tickers, weights):
        rows.append({"Stock": ticker, "Weight (%)": round(w * 100, 2)})

    df = pd.DataFrame(rows)
    df.loc[len(df)] = {"Stock": "", "Weight (%)": ""}           # blank spacer
    df.loc[len(df)] = {"Stock": "Portfolio Return (%)",  "Weight (%)": round(ret * 100, 2)}
    df.loc[len(df)] = {"Stock": "Portfolio Volatility (%)", "Weight (%)": round(vol * 100, 2)}
    df.loc[len(df)] = {"Stock": "Portfolio Sharpe Ratio",   "Weight (%)": round(sharpe, 4)}
    return df, ret, vol, sharpe


def print_portfolio(label, tickers, weights, mu, cov):
    ret, vol, sharpe = portfolio_stats(weights, mu, cov)
    width = 62
    print(f"\n{'='*width}")
    print(f"  {label}")
    print(f"{'='*width}")
    print(f"  {'Stock':<20}  {'Weight':>10}")
    print(f"  {'-'*20}  {'-'*10}")
    for ticker, w in zip(tickers, weights):
        if w >= 0.0001:   # hide dust-level allocations
            print(f"  {ticker:<20}  {w*100:>9.2f}%")
    print(f"  {'-'*20}  {'-'*10}")
    print(f"\n  Expected Annual Return : {ret*100:>8.2f}%")
    print(f"  Annual Volatility      : {vol*100:>8.2f}%")
    print(f"  Sharpe Ratio           : {sharpe:>8.4f}")
    print(f"{'='*width}")


def export_to_excel(portfolios, out_path):
    """
    portfolios: list of (sheet_name, result_df) tuples.
    Writes each to its own sheet in out_path.
    """
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, df in portfolios:
            # Keep only weight rows (exclude spacer and summary rows) for
            # the stock block, then append a stats block below
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]
            # Bold the header row
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill("solid", fgColor="1F4E79")
            for cell in ws[1]:
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Right-align the Weight column; highlight summary rows
            summary_fill = PatternFill("solid", fgColor="D9E1F2")
            for row in ws.iter_rows(min_row=2):
                label_cell  = row[0]
                weight_cell = row[1]
                weight_cell.alignment = Alignment(horizontal="right")
                if label_cell.value in (
                    "Portfolio Return (%)",
                    "Portfolio Volatility (%)",
                    "Portfolio Sharpe Ratio",
                ):
                    for cell in row:
                        cell.font = Font(bold=True)
                        cell.fill = summary_fill

            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 16


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print(f"\n{'='*62}")
    print("  Module 2 — Layer 3 Core Optimizer")
    print("  Goals: Min Variance (GMV) | Max Risk-Adjusted | Tail-Risk (stub)")
    print(f"{'='*62}\n")

    # ── 1. Load pre-computed mu and covariance from module1 ───────────────────
    returns_path = os.path.join(SCRIPT_DIR, "returns_stats.xlsx")
    if not os.path.exists(returns_path):
        print(f"  ERROR: '{returns_path}' not found.")
        print("  Run module1_data.py first to generate the required files.")
        sys.exit(1)

    mu  = None
    cov = None
    tickers = None

    # Primary path: read pre-computed split-estimation mu and cov
    try:
        mu_df  = pd.read_excel(returns_path, sheet_name="Annualised Mu")
        cov_df = pd.read_excel(returns_path, sheet_name="Annualised Cov", index_col=0)
        tickers = list(mu_df["Ticker"])
        mu      = mu_df["Annualised_Expected_Return"].values.astype(float)
        cov     = cov_df.loc[tickers, tickers].values.astype(float)
        print(f"  Loaded split-estimation parameters for {len(tickers)} stocks")
        print( "  Return source : momentum 12M-1M  (Annualised Mu sheet)")
        print( "  Risk source   : 10-year monthly  (Annualised Cov sheet)")
    except Exception:
        pass

    # Fallback: recompute from daily returns if new sheets are absent
    if mu is None:
        print("  WARNING: 'Annualised Mu'/'Annualised Cov' sheets not found.")
        print("           Re-run module1_data.py to generate split-estimation parameters.")
        print("           Falling back to daily-returns-based estimation.\n")
        try:
            returns_df = pd.read_excel(
                returns_path, sheet_name="Daily Returns", index_col=0, parse_dates=True
            )
        except Exception as exc:
            print(f"  ERROR reading returns_stats.xlsx: {exc}")
            sys.exit(1)
        returns_df = returns_df.dropna(how="all").dropna(axis=1, how="all")
        if returns_df.empty:
            print("  ERROR: 'Daily Returns' sheet is empty.")
            sys.exit(1)
        tickers = list(returns_df.columns)
        clean   = returns_df.dropna()
        mu      = clean.mean().values * TRADING_DAYS
        cov     = clean.cov().values  * TRADING_DAYS

    n = len(tickers)
    print(f"\n  Annualised expected returns (%):")
    for t, r in zip(tickers, mu):
        print(f"    {t:<22} {r*100:>7.2f}%")

    # ── 3. Feasibility check (for goal 2's adaptive cap) ──────────────────────
    _, w_cap = get_adaptive_bounds(n)
    min_stocks = int(np.ceil(1.0 / w_cap))
    if n < min_stocks:
        print(f"\n  ERROR: Need at least {min_stocks} stocks for the "
              f"{w_cap*100:.0f}% cap (adaptive, n={n}). Only {n} available.")
        sys.exit(1)
    print(f"\n  Goal 1 (GMV) bounds : 0%..100% (no caps)")
    print(f"  Goal 2 cap (n={n})    : 0%..{w_cap*100:.0f}% (adaptive)")

    # ── 3b. Load FACTOR returns for the Max Risk-Adjusted goal ────────────────
    #        (FF3-US / Carhart4-India model, written by module 1 as "CAPM Returns".)
    capm_mu = mu.copy()   # default: fall back to momentum if factor mu unavailable
    try:
        capm_mu = load_capm_mu(returns_path, tickers, mu)
        print(f"  Factor returns loaded for the Max Risk-Adjusted goal.")
        for t, r in zip(tickers, capm_mu):
            print(f"    {t:<22} {r*100:>7.2f}%  (factor)")
    except Exception as exc:
        print(f"  WARNING: Could not load CAPM Returns sheet ({exc}). "
              "Max Risk-Adjusted will fall back to momentum returns.")

    # ── 4. Run the three Layer 3 goals ─────────────────────────────────────────
    portfolios = []   # (sheet_name, result_df)

    # Goal 1 — Minimum Variance (GMV): minimise wᵀΣw, w>=0, sum=1, NO caps.
    #          Objective ignores returns; momentum mu is passed for display only.
    label_gmv = "1 — Minimum Variance (GMV)"
    print(f"\n  Optimising: {label_gmv} ...  (no concentration caps)", end=" ", flush=True)
    w_gmv, ok_gmv, msg_gmv = optimise(portfolio_vol, mu, cov, label_gmv,
                                      w_min=0.0, w_max=1.0)
    if ok_gmv:
        print("done.")
        print_portfolio(label_gmv, tickers, w_gmv, mu, cov)
        df_gmv, _, _, _ = build_result_df(tickers, w_gmv, mu, cov)
        portfolios.append(("Minimum Variance", df_gmv))
    else:
        print(f"\n  WARNING: {msg_gmv}")

    # Goal 2 — Maximum Risk-Adjusted Return: maximise Sharpe with the FACTOR mu
    #          (capm_mu), same Σ, rf=blended_rate, PLUS adaptive caps (no min floor).
    adaptive_cap = get_adaptive_bounds(n)[1]   # max only: 60% / 35% / 20%
    label_rar = "2 — Maximum Risk-Adjusted Return"
    print(f"\n  Optimising: {label_rar} ...  (adaptive cap {adaptive_cap*100:.0f}%, "
          f"factor mu)", end=" ", flush=True)
    w_rar, ok_rar, msg_rar = optimise(neg_sharpe, capm_mu, cov, label_rar,
                                      w_min=0.0, w_max=adaptive_cap)
    if ok_rar:
        print("done.")
        print_portfolio(label_rar, tickers, w_rar, capm_mu, cov)
        max_w = float(np.max(w_rar))
        print(f"  Cap check              : max weight {max_w*100:.2f}% "
              f"<= {adaptive_cap*100:.0f}%  -> {'OK' if max_w <= adaptive_cap + 1e-6 else 'VIOLATED'}")
        df_rar, _, _, _ = build_result_df(tickers, w_rar, capm_mu, cov)
        portfolios.append(("Max Risk-Adjusted", df_rar))
    else:
        print(f"\n  WARNING: {msg_rar}")

    # Goal 3 — Tail-Risk Minimization (Layer 6 CVaR optimizer over Layer 4 scenarios).
    #          Reads simulated_returns.npz; minimises CVaR(95%) with a (default
    #          non-binding) annualised return floor. Never fabricates weights:
    #          a missing scenario file or an infeasible floor surfaces a message.
    sim_path = os.path.join(SCRIPT_DIR, "simulated_returns.npz")
    if not os.path.exists(sim_path):
        print(f"\n{'='*62}")
        print("  3 — Tail-Risk Minimization (CVaR 95%)  [Layer 6]")
        print(f"{'='*62}")
        print("  Layer 4 scenarios not found (simulated_returns.npz).")
        print("  Run simulation_engine.py first; skipping (no weights produced).")
        print(f"{'='*62}")
    else:
        scen, scen_tickers = tro.load_scenarios(sim_path)
        tail_result = tro.solve_min_cvar(scen, scen_tickers, r_min=0.0)
        tro.print_tail_risk_result(tail_result)
        if tail_result["status"] == "optimal":
            portfolios.append(("Tail-Risk CVaR", tro.build_export_df(tail_result)))

    if not portfolios:
        print("\n  All optimisations failed. No output file written.")
        sys.exit(1)

    # ── 5. Export to Excel ─────────────────────────────────────────────────────
    out_path = os.path.join(SCRIPT_DIR, "optimised_portfolios.xlsx")
    try:
        export_to_excel(portfolios, out_path)
        print(f"\n  Exported -> {out_path}")
    except Exception as exc:
        print(f"\n  ERROR writing optimised_portfolios.xlsx: {exc}")
        sys.exit(1)

    print(f"\n{'='*62}\n")


if __name__ == "__main__":
    main()
