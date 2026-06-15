"""
Asian currencies vs USD — one-year change (rolling: today − 1 year → today)
"""

import yfinance as yf
import pandas as pd
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────

CURRENCIES = {
    "JPY": ("Japanese Yen",       "USDJPY=X"),
    "CNY": ("Chinese Yuan",       "USDCNY=X"),
    "INR": ("Indian Rupee",       "USDINR=X"),
    "KRW": ("Korean Won",         "USDKRW=X"),
    "HKD": ("Hong Kong Dollar",   "USDHKD=X"),
    "SGD": ("Singapore Dollar",   "USDSGD=X"),
    "TWD": ("Taiwan Dollar",      "USDTWD=X"),
    "THB": ("Thai Baht",          "USDTHB=X"),
    "MYR": ("Malaysian Ringgit",  "USDMYR=X"),
    "IDR": ("Indonesian Rupiah",  "USDIDR=X"),
    "PHP": ("Philippine Peso",    "USDPHP=X"),
    "VND": ("Vietnamese Dong",    "USDVND=X"),
}

END_DATE   = date.today()
START_DATE = END_DATE.replace(year=END_DATE.year - 1)
OUTPUT     = r"C:\Users\adity\Documents\FinanceScripts\asian_fx_vs_usd.xlsx"

# ── Data retrieval ─────────────────────────────────────────────────────────────

def nearest_value(series: pd.Series, target: date, window: int = 5) -> float | None:
    """Return the closing price on or nearest before `target` within `window` days."""
    for delta in range(window + 1):
        d = pd.Timestamp(target - timedelta(days=delta))
        if d in series.index:
            return series[d]
    return None


def fetch_rates() -> pd.DataFrame:
    tickers = [info[1] for info in CURRENCIES.values()]
    # Fetch a small buffer beyond our dates to handle weekends/holidays
    raw = yf.download(
        tickers,
        start=str(START_DATE - timedelta(days=7)),
        end=str(END_DATE + timedelta(days=3)),
        auto_adjust=True,
        progress=False,
    )["Close"]

    records = []
    for code, (name, ticker) in CURRENCIES.items():
        series = raw[ticker].dropna() if ticker in raw.columns else pd.Series(dtype=float)
        rate_start = nearest_value(series, START_DATE)
        rate_end   = nearest_value(series, END_DATE)

        if rate_start is None or rate_end is None:
            pct_change = None
            direction  = "N/A"
        else:
            pct_change = (rate_end - rate_start) / rate_start * 100
            # A higher USD/XXX rate means the local currency weakened vs USD
            direction = "Weakened" if pct_change > 0 else ("Strengthened" if pct_change < 0 else "Unchanged")

        records.append({
            "Currency Code":        code,
            "Currency Name":        name,
            "Ticker":               ticker,
            f"Rate {START_DATE}":   round(rate_start, 4) if rate_start else None,
            f"Rate {END_DATE}":     round(rate_end,   4) if rate_end   else None,
            "1Y Change (%)":        round(pct_change, 2) if pct_change is not None else None,
            "Direction vs USD":     direction,
        })

    return pd.DataFrame(records)


# ── Excel export ───────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
GREEN_FILL   = PatternFill("solid", fgColor="C6EFCE")
RED_FILL     = PatternFill("solid", fgColor="FFC7CE")
NEUTRAL_FILL = PatternFill("solid", fgColor="FFEB9C")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def style_sheet(ws, df: pd.DataFrame):
    # Header row
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill       = HEADER_FILL
        cell.font       = HEADER_FONT
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border     = THIN_BORDER

    # Data rows
    pct_col = df.columns.get_loc("1Y Change (%)") + 1
    dir_col = df.columns.get_loc("Direction vs USD") + 1

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        pct = row._asdict().get("1Y Change (%)")
        direction = row._asdict().get("Direction vs USD")

        if direction == "Strengthened":
            row_fill = GREEN_FILL
        elif direction == "Weakened":
            row_fill = RED_FILL
        else:
            row_fill = NEUTRAL_FILL

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if col_idx in (pct_col, dir_col):
                cell.fill = row_fill
            if col_idx == pct_col and isinstance(value, (int, float)):
                cell.number_format = "+0.00%;-0.00%;0.00%"
                cell.value = value / 100  # store as fraction for % format

    # Column widths
    col_widths = [10, 22, 14, 16, 16, 16, 18]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"


def add_summary(wb, df: pd.DataFrame):
    ws = wb.create_sheet("Summary")
    valid = df.dropna(subset=["1Y Change (%)"])
    strongest = valid.loc[valid["1Y Change (%)"].idxmin()]
    weakest   = valid.loc[valid["1Y Change (%)"].idxmax()]

    rows = [
        ("Period",          f"{START_DATE}  →  {END_DATE}"),
        ("Currencies",      len(df)),
        ("Strongest vs USD", f"{strongest['Currency Code']} ({strongest['1Y Change (%)']:+.2f}%)"),
        ("Weakest vs USD",  f"{weakest['Currency Code']} ({weakest['1Y Change (%)']:+.2f}%)"),
        ("Avg Change (%)",  f"{valid['1Y Change (%)'].mean():+.2f}%"),
    ]

    for r, (label, value) in enumerate(rows, start=1):
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=value)
        lc.font = Font(bold=True)
        lc.border = vc.border = THIN_BORDER
        lc.alignment = vc.alignment = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30


def export_excel(df: pd.DataFrame, path: str):
    df.to_excel(path, index=False, sheet_name="FX Rates")
    wb = load_workbook(path)
    ws = wb["FX Rates"]
    style_sheet(ws, df)
    add_summary(wb, df)
    wb.save(path)
    print(f"Saved → {path}")


# ── Main ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"Fetching exchange rates for {len(CURRENCIES)} Asian currencies …")
    df = fetch_rates()

    print("\n── Results ──────────────────────────────────────────────────────")
    print(df.to_string(index=False))
    print()

    export_excel(df, OUTPUT)
